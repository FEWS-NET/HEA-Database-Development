# Will be baseline.pipelines.ingestion
# @TODO Review imports
# import datetime
# import hashlib
import io
import itertools
import logging
import re
from collections.abc import Iterable
from pathlib import Path

import luigi
import openpyxl
import pandas as pd
import xlrd
import xlwt
from django.core.management import call_command
from django.db.models import ForeignKey
from kiluigi import PipelineError
from kiluigi.format import Pickle
from kiluigi.targets import GoogleDriveTarget, IntermediateTarget, LocalTarget
from kiluigi.utils import class_from_name, path_from_task

# from luigi.task import TASK_ID_INVALID_CHAR_REGEX, TASK_ID_TRUNCATE_HASH
from luigi.util import requires
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, rows_from_range
from xlutils.copy import copy as copy_xls

from baseline.models import WealthGroupCharacteristicValue
from common.lookups import ClassifiedProductLookup, UnitOfMeasureLookup
from common.management.commands import verbose_load_data
from common.pipelines.format import JSON
from metadata.lookups import WealthCharacteristicLookup, WealthGroupCategoryLookup
from metadata.models import WealthCharacteristic

# from slugify import slugify


logger = logging.getLogger(__name__)


def get_index(search_text: str | list[str], data: pd.Series) -> tuple[int, int]:
    """
    Return the index of the first value in a Series that matches the text.

    Note that the search is case-insensitive.

    The text to search for can be a string or a list of strings, in which case the function
    returns the first cell that matches any of the supplied strings.
    """
    # Make sure we have an iterable that we can pass to `.isin()`
    if isinstance(search_text, str) or not isinstance(search_text, Iterable):
        search_text = [str(search_text)]
    # Make sure that the search terms are lowercase
    search_text = [str(search_term).lower() for search_term in search_text]
    # Convert the Series to a set of True/False values based on whether they match one of the
    # search_text values, and use idxmax to return the index of the first match.
    # This works because in Pandas True > False, so idxmax() returns the index of the first True.
    index = data.str.lower().isin(search_text).idxmax()
    return index


def get_unmatched_metadata(df: pd.DataFrame, column: str) -> pd.Series:
    """
    Return a series of original values that were not matched by a Lookup.
    """
    unmatched_metadata = (
        df[(df[column].isnull()) & ~((df[column + "_original"].isnull()) | (df[column + "_original"].eq("")))][
            column + "_original"
        ]
        .sort_values()
        .unique()
    )
    return unmatched_metadata


def verbose_pivot(df: pd.DataFrame, values: str | list[str], index: str | list[str], columns: str | list[str]):
    """
    Pivot a DataFrame, or log a detailed exception in the event of a failure

    Failures are typically caused by duplicate entries in the index.
    """
    # Make sure index and columns are lists so we can concatenate them in the error handler, if needed.
    if isinstance(index, str):
        index = [index]
    if isinstance(columns, str):
        columns = [columns]
    try:
        return pd.pivot(df, values=values, index=index, columns=columns).reset_index()
    except ValueError as e:
        # Need to fillna, otherwise the groupby returns an empty dataframe
        duplicates = df.fillna("").groupby(index + columns).size().reset_index(name="count")

        # Filter for the entries that appear more than once, i.e., duplicates
        duplicates = duplicates[duplicates["count"] > 1]

        error_df = pd.merge(df.fillna(""), duplicates[index + columns], on=index + columns)

        raise PipelineError(str(e) + "\n" + error_df.to_markdown()) from e


class GetBSS(luigi.ExternalTask):
    """
    External Task that returns a Target for accessing the BSS spreadsheet.
    """

    bss_path = luigi.Parameter(description="Path to the BSS file")

    def output(self):
        target = LocalTarget(Path(self.bss_path).expanduser().absolute(), format=luigi.format.Nop)
        if not target.exists():
            target = GoogleDriveTarget(
                self.bss_path,
                format=luigi.format.Nop,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            if not target.exists():
                raise PipelineError("No local or Google Drive file matching %s" % self.bss_path)
        return target


class GetBSSMetadata(luigi.ExternalTask):
    """
    External Task that returns a spreadsheet of metadata about all BSSs.
    """

    metadata_path = luigi.Parameter(description="Path to the BSS metadata")

    def output(self):
        target = LocalTarget(Path(self.metadata_path).expanduser().absolute(), format=luigi.format.Nop)
        if not target.exists():
            target = GoogleDriveTarget(
                self.metadata_path,
                format=luigi.format.Nop,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            if not target.exists():
                raise PipelineError("No local or Google Drive file matching %s" % self.metadata_path)
        return target


class GetBSSCorrections(luigi.Task):
    """
    Return a DataFrame of corrections to make to one or more BSSs.
    """

    corrections_path = luigi.Parameter(default="", description="Path to the BSS corrections")

    def output(self):
        return IntermediateTarget(path_from_task(self) + ".pickle", format=Pickle, timeout=3600)

    def run(self):
        if self.corrections_path:
            target = LocalTarget(Path(self.corrections_path).expanduser().absolute(), format=luigi.format.Nop)
            if not target.exists():
                target = GoogleDriveTarget(
                    self.corrections_path,
                    format=luigi.format.Nop,
                    mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
                if not target.exists():
                    raise PipelineError("No local or Google Drive file matching %s" % self.corrections_path)
            with target.open() as input:
                corrections_df = pd.read_excel(input, "Corrections")
        else:
            corrections_df = pd.DataFrame(
                columns=[
                    "bss_path",
                    "worksheet_name",
                    "range",
                    "prev_value",
                    "value",
                    "correction_date",
                    "author",
                    "comment",
                ]
            )

        with self.output().open("w") as output:
            output.write(corrections_df)


@requires(bss=GetBSS, corrections=GetBSSCorrections)
class CorrectBSS(luigi.Task):
    def output(self):
        return IntermediateTarget(
            path_from_task(self) + Path(self.input()["bss"].path).suffix, format=luigi.format.Nop, timeout=3600
        )

    def run(self):
        with self.input()["corrections"].open() as input:
            corrections_df = input.read()

        # Find the <country>/<filename>.<ext> path to the file, relative to the root folder.
        path = f"{Path(self.bss_path).parent.name}/{Path(self.bss_path).name}"
        # Find the corrections for this BSS
        corrections_df = corrections_df[corrections_df["bss_path"] == path]

        if corrections_df.empty:
            # No corrections, so just leave the file unaltered
            with self.output().open("w") as output, self.input()["bss"].open() as input:
                output.write(input.read())

        else:
            if Path(self.input()["bss"].path).suffix == ".xls":
                # xlrd can only read XLS files, so we need to use xlutils.copy to create something we can edit
                with self.input()["bss"].open() as input:
                    wb = xlrd.open_workbook(file_contents=input.read(), formatting_info=True, on_demand=True)
            else:
                with self.input()["bss"].open() as input:
                    wb = openpyxl.load_workbook(input)
            wb = self.process(wb, corrections_df)
            with self.output().open("w") as output:
                wb.save(output)

    def process(
        self,
        wb: xlrd.book.Book | openpyxl.Workbook,
        corrections_df: pd.DataFrame,
    ) -> xlwt.Workbook | openpyxl.Workbook:
        """
        Process the Excel workbook and apply corrections and then return the corrected file.
        """
        if isinstance(wb, xlrd.book.Book):
            xlrd_wb = wb
            wb = copy_xls(wb)  # a writable workbook to be returned by this method
        else:
            xlrd_wb = None  # Required to suppress spurious unbound variable errors from Pyright
        for correction in corrections_df.itertuples():
            for row in rows_from_range(correction.range):
                for cell in row:
                    if isinstance(wb, xlwt.Workbook):
                        row, col = coordinate_to_tuple(cell)
                        prev_value = xlrd_wb.sheet_by_name(correction.worksheet_name).cell_value(row - 1, col - 1)
                        if (
                            xlrd_wb.sheet_by_name(correction.worksheet_name).cell(row - 1, col - 1).ctype
                            == xlrd.XL_CELL_ERROR
                        ):
                            # xlrd.error_text_from_code returns, eg, "#N/A"
                            prev_value = xlrd.error_text_from_code[
                                xlrd_wb.sheet_by_name(correction.worksheet_name).cell_value(row - 1, col - 1)
                            ]
                        self.validate_previous_value(cell, correction.prev_value, prev_value)
                        # xlwt uses 0-based indexes, but coordinate_to_tuple uses 1-based, so offset the values
                        wb.get_sheet(correction.worksheet_name).write(row - 1, col - 1, correction.value)
                    else:
                        cell = wb[correction.worksheet_name][cell]
                        self.validate_previous_value(cell, correction.prev_value, cell.value)
                        cell.value = correction.value
                        cell.comment = Comment(
                            f"{correction.author} on {correction.correction_date.date().isoformat()}: {correction.comment}",  # NOQA: E501
                            author=correction.author,
                        )

        return wb

    def validate_previous_value(self, cell, expected_prev_value, prev_value):
        # "#N/A" is inconsistently loaded as nan, even when copied and pasted in Excel or GSheets
        if (str(expected_prev_value) != str(prev_value)) & (
            {"nan", "#N/A"} != {str(expected_prev_value), str(prev_value)}
        ):
            raise PipelineError(
                "Unexpected prior value in source BSS. "
                f"BSS `{self.input()['bss'].name}`, cell `{cell}`, "
                f"value found `{prev_value}`, expected `{expected_prev_value}`."
            )


@requires(bss=CorrectBSS, metadata=GetBSSMetadata)
class NormalizeWB(luigi.Task):
    def output(self):
        return IntermediateTarget(path_from_task(self) + ".json", format=JSON, timeout=3600)

    def run(self):
        with self.input()["bss"].open() as input:
            data_df = pd.read_excel(input, "WB", header=None)
        with self.input()["metadata"].open() as input:
            metadata_df = pd.read_excel(input, "Metadata")

        # Find the <country>/<filename>.<ext> path to the file, relative to the root folder.
        path = f"{Path(self.bss_path).parent.name}/{Path(self.bss_path).name}"
        # Find the metadata for this BSS
        try:
            metadata = metadata_df[metadata_df["bss_path"] == path].iloc[0]
        except IndexError:
            raise PipelineError("No entry in BSS Metadata worksheet for %s" % path)

        data = self.process(data_df, metadata)

        with self.output().open("w") as output:
            output.write(data)

    def process(self, data_df: pd.DataFrame, metadata: pd.Series) -> dict:
        """
        Process the WB Sheet and return a Python dict of normalized data.

        The first rows of the sheet look like:
        | Row | A                         | B                   | C                                          | D                                          | E                                          | F                                          | G                |
        |-----|---------------------------|---------------------|--------------------------------------------|--------------------------------------------|--------------------------------------------|--------------------------------------------|------------------|
        |   0 | MALAWI HEA BASELINES 2015 | Southern Lakeshore  | Southern Lakeshore                         |                                            |                                            |                                            |                  |
        |   1 |                           |                     | Community interviews                       |                                            |                                            |                                            |                  |
        |   2 | WEALTH GROUP              |                     |                                            |                                            |                                            |                                            |                  |
        |   3 | District                  | Salima and Mangochi | Salima                                     | Salima                                     | Salima                                     | Salima                                     | Dedza            |
        |   4 | Village                   |                     | Mtika                                      | Pemba                                      | Ndembo                                     | Makanjira                                  | Kasakala         |
        |   5 | Interview number:         |                     | 1                                          | 2                                          | 3                                          | 4                                          | 5                |
        |   6 | Interviewers              |                     | Kandiwo, Ethel, Fyawupi, Daniel, Chipiliro | Kandiwo, Ethel, Fyawupi, Daniel, Chipiliro | Kandiwo, Ethel, Fyawupi, Chipiliro, Daniel | Kandiwo, Ethel, Fyawupi, Chipiliro, Daniel | Chipiliro, Imran |
        """  # NOQA: E501
        # Key in the result dict must match the name of the model class they will be imported to.
        # The value must be a list of instances to import into that model, where each instance
        # is a dict of field names and values.
        # If the field is a foreign key to a model that supports a natural key (i.e. the model has a `natural_key`
        # method), then the field value should be a list of components to the natural key.
        result = {
            "LivelihoodZone": [
                {
                    # Get country and code from the filename
                    "country": metadata["country"],
                    "code": metadata["code"],
                    "name": metadata["name"],
                }
            ],
            "LivelihoodZoneBaseline": [
                {
                    "livelihood_zone": metadata["code"],
                    "name": metadata["name"],
                    "source_organization": [
                        metadata["source_organization"],
                    ],  # natural key is always a list
                    "main_livelihood_category": str(metadata["main_livelihood_category"]).lower(),
                    "reference_year_start_date": metadata["reference_year_start_date"].date().isoformat(),
                    "reference_year_end_date": metadata["reference_year_end_date"].date().isoformat(),
                    "valid_from_date": metadata["valid_from_date"].date().isoformat(),
                    "valid_to_date": None
                    if pd.isna(metadata["valid_to_date"])
                    else metadata["valid_to_date"].date().isoformat(),
                    "data_collection_start_date": None
                    if pd.isna(metadata["data_collection_start_date"])
                    else metadata["data_collection_start_date"].date().isoformat(),
                    "data_collection_end_date": None
                    if pd.isna(metadata["data_collection_end_date"])
                    else metadata["data_collection_end_date"].date().isoformat(),
                    "publication_date": None
                    if pd.isna(metadata["publication_date"])
                    else metadata["publication_date"].date().isoformat(),
                }
            ],
        }

        # Find the communities
        # Transpose to get data in columns
        community_df = data_df.iloc[2:7].transpose()
        # Check that the columns are what we expect
        expected_column_sets = (["WEALTH GROUP", "District", "Village", "Interview number:", "Interviewers"],)
        found_columns = community_df.iloc[0].str.strip().tolist()
        if not any(found_columns == expected_columns for expected_columns in expected_column_sets):
            raise PipelineError("Cannot identify Communities from columns %s" % ", ".join(found_columns))
        # Normalize the column names
        community_df.columns = ["wealth_group_category", "district", "name", "interview_number", "interviewers"]
        community_df = community_df[1:]
        # Find the initial set of Communities by only finding rows that have both a `district` and a `community`,
        # and don't have a `wealth_group_category`. Also ignore the `Comments` row.
        community_df = (
            community_df[(community_df["district"] != "Comments") & (community_df["wealth_group_category"].isna())][
                ["district", "name", "interview_number", "interviewers"]
            ]
            .dropna(subset=["name"])
            .drop_duplicates()
        )
        # Create the full_name from the community and district, and fall back
        # to just the community name if the district is empty/nan.
        community_df["full_name"] = community_df.name.str.cat(community_df.district, sep=", ").fillna(
            community_df.name
        )
        community_df = community_df.drop(columns="district")
        # Add the natural key for the livelihood zone baseline
        community_df["livelihood_zone_baseline"] = community_df["full_name"].apply(
            lambda full_name: [metadata["code"], metadata["reference_year_end_date"].date().isoformat()]
        )

        # Replace NaN with "" ready for Django
        community_df = community_df.fillna("")
        result["Community"] = community_df.to_dict(orient="records")

        # Process the main part of the sheet to find the Wealth Group Characteristic Values

        # Find the column index of the last column that we are contains relevant data.
        # The final three relevant columns are Summary/From/To. Find Summary because
        # it is less likely to give a false positive than the From or To, and then add 2.
        last_col_index = get_index(["Summary"], data_df.iloc[3]) + 2

        # Create charactistics dataframe containing data from row 9 onwards and from column A
        # up to the summary/from/to columns
        char_df = data_df.iloc[8:, : last_col_index + 1]

        # Keep the row number to aid trouble-shooting.
        char_df["row"] = char_df.index + 1

        # Set the column names based on the interviewee wealth group category (Row 3)
        # and community full name (Rows 4 & 5)
        row_3_values = data_df.iloc[2, 2 : last_col_index - 2].values
        row_4_values = data_df.iloc[3, 2 : last_col_index - 2].values
        row_5_values = data_df.iloc[4, 2 : last_col_index - 2].values
        new_column_names = []
        for wealth_group_category, admin_name, community_name in zip(row_3_values, row_4_values, row_5_values):
            full_name = f"{community_name}, {admin_name}" if pd.notna(admin_name) else community_name
            wealth_group_category = wealth_group_category if pd.notna(wealth_group_category) else ""
            new_column_names.append(f"{full_name}:{wealth_group_category}")
        new_column_names = (
            ["characteristic_label", "wealth_group_category"]
            + new_column_names
            + ["summary", "min_value", "max_value", "row"]
        )
        char_df.columns = new_column_names

        # Fill NaN values in 'wealth_characteristic' from the previous non-null cell
        char_df["characteristic_label"] = char_df["characteristic_label"].ffill()

        # Get the 'wealth_characteristic', 'product' and 'unit_of_measure' from the 'characteristic_label'
        # We can't just use a regular WealthCharacteristicLookup.do_lookup() because we need to match the additional
        # 'product' and 'unit_of_measure' attributes using regexes. Therefore, we use regexes against the aliases to
        # set the product and/or unit_of_measure for those wealth characteristics where we expect those attributes to
        # be present, and then we use the regular WealthCharacteristicLookup.do_lookup() to match the remaining items.

        # Examples:
        # MWPHA_30Sep15: Land area cultivated - rainfed crops (acres)
        # NE01(BIL): Anes nbr possédés
        # NE01(BIL): Volaille nbr possédés
        # LR02: Oxen (no. owned)
        # SO18: Donkey number owned

        # Build a map of regex strings to the actual characteristic, using the WealthCharacteristic.aliases
        wealth_characteristic_map = {}
        for wealth_characteristic in WealthCharacteristic.objects.exclude(
            has_product=False, has_unit_of_measure=False
        ).exclude(aliases__isnull=True):
            for alias in wealth_characteristic.aliases:
                alias = re.escape(alias)
                # replace the <product> or <unit_of_measure> placeholders with regex groups
                for attribute in "product", "unit_of_measure":
                    if attribute in alias:
                        alias = alias.replace(f"<{attribute}>", f"(?P<{attribute}>[a-z][a-z']+)")
                alias = re.compile(alias)
                wealth_characteristic_map[alias] = wealth_characteristic.code

        def get_attributes(characteristic_label: str):
            """
            Return a tuple of (wealth_characteristic, product, unit_of_measure) from a characteristic label.
            """
            for pattern, wealth_characteristic in wealth_characteristic_map.items():
                match = pattern.fullmatch(characteristic_label.lower().strip())
                if match:
                    if "product" in match.groupdict():
                        product = match.groupdict()["product"]
                    else:
                        product = None
                    if "unit_of_measure" in match.groupdict():
                        unit_of_measure = match.groupdict()["unit_of_measure"]
                    else:
                        unit_of_measure = None
                    # Return a Series so that Pandas will split the return value into multiple columns
                    return pd.Series([wealth_characteristic, product, unit_of_measure])
            # No pattern matched
            # Return a Series so that Pandas will split the return value into multiple columns
            return pd.Series([None, None, None])

        # Extract the wealth_characteristic, product and unit_of_measure (if applicable) from the characteristic_label
        char_df[["wealth_characteristic", "product", "unit_of_measure"]] = char_df["characteristic_label"].apply(
            get_attributes
        )

        # Look up the metadata codes for WealthGroupCategory and WealthCharacteristic
        # We need to lookup the WealthCharacteristic before we can use it to mask the 'product' column below
        char_df = WealthGroupCategoryLookup().do_lookup(char_df, "wealth_group_category", "wealth_group_category")
        # Note that we use do_lookup(update=True) so that we ignore the values that we matched using a regex
        char_df = WealthCharacteristicLookup().do_lookup(
            char_df, "characteristic_label", "wealth_characteristic", update=True
        )

        # Report any errors, because unmatched WealthCharacteristics that contain a product reference may cause errors
        # when reshaping the dataframe. Note that we don't raise an exception here - ValidateBaseline will do that if
        # we get that far. But we do print the error message to help with troubleshooting.
        for value in get_unmatched_metadata(char_df, "wealth_characteristic"):
            logger.warning(
                f"Unmatched remote metadata for WealthGroupCharacteristicValue with wealth_characteristic '{value}'"
            )  # NOQA: E501

        # Copy the product from the previous cell for rows that need a product but don't specify it.
        # We do this by setting the missing values to pd.NA and then using .ffill()
        # Note that we need to replace the None with something else during the ffill() so that it is only actual pd.NA
        # values that are replaced
        product_characteristics = WealthCharacteristic.objects.filter(has_product=True).values_list("code", flat=True)
        char_df["product"] = (
            char_df["product"]
            .replace({None: ""})
            .mask(
                char_df["wealth_characteristic"].isin(list(product_characteristics)) & pd.isna(char_df["product"]),
                pd.NA,
            )
            .ffill()
            .replace({"": None})
        )

        # Copy the unit_of_measure from the previous cell for rows that need a unit_of_measure but don't specify it.
        # We do this by setting the missing values to pd.NA and then using .ffill()
        # Note that we need to replace the None with something else during the ffill() so that it is only actual pd.NA
        # values that are replaced
        unit_of_measure_characteristics = WealthCharacteristic.objects.filter(has_unit_of_measure=True).values_list(
            "code", flat=True
        )
        char_df["unit_of_measure"] = (
            char_df["unit_of_measure"]
            .replace({None: ""})
            .mask(
                char_df["wealth_characteristic"].isin(list(unit_of_measure_characteristics))
                & pd.isna(char_df["unit_of_measure"]),
                pd.NA,
            )
            .ffill()
            .replace({"": None})
        )

        # Lookup the CPCv2
        char_df = ClassifiedProductLookup().do_lookup(char_df, "product", "product")

        # Lookup the Unit of Measure
        char_df = UnitOfMeasureLookup().do_lookup(char_df, "unit_of_measure", "unit_of_measure")

        # Melt the dataframe
        char_df = pd.melt(
            char_df,
            id_vars=[
                "row",
                "characteristic_label",
                "product_original",
                "product",
                "unit_of_measure_original",
                "unit_of_measure",
                "wealth_characteristic_original",
                "wealth_characteristic",
                "wealth_group_category_original",
                "wealth_group_category",
            ],
            value_vars=char_df.columns[2:-1],
            var_name="interview_label",
            value_name="value",
        )

        # Add a column containing the original column label, to aid trouble-shooting.
        mapping_dict = {value: get_column_letter(idx + 1) for idx, value in enumerate(new_column_names)}
        mapping_dict["row"] = None  # column 'row' was added by us and isn't a column in the workbook
        char_df["column"] = char_df["interview_label"].map(mapping_dict)

        # Split 'label' into 'interviewee_wealth_group_category' and 'full_name'
        char_df[["full_name", "interviewee_wealth_group_category"]] = char_df["interview_label"].str.split(
            ":", expand=True, n=1
        )

        # Drop rows where 'value' is NaN,
        # or wealth category is blank (which is the total row for percentage of households)
        char_df = char_df.dropna(subset=["value", "wealth_group_category"])

        # Also drop rows where there is no Community, and which aren't the summary.
        # These rows came from blank columns in the worksheet.from django.core.exceptions import ValidationError

        char_df = char_df[~char_df["interview_label"].eq("nan:")]

        # Drop rows where 'value' is NaN,
        # or wealth group category is blank (which is the total row for percentage of households)
        char_df = char_df.dropna(subset=["value", "wealth_group_category"])

        # Also drop rows containing the percentage of households estimate from a Wealth Group-level interview,
        # estimating the percentage for the other Wealth Categories. I.e. for the "VP" Wealth Group interview
        # we only store the percentage of households estimated for the "VP" Wealth Group Category, and not
        # those for P, M and B/O.
        char_df = char_df[
            ~(
                char_df["wealth_characteristic"].eq("percentage of households")
                & ~char_df["interviewee_wealth_group_category"].isna()
                & ~char_df["wealth_group_category_original"].eq(char_df["interviewee_wealth_group_category"])
            )
        ]

        # Split out the summary data
        summary_df = char_df[char_df["interview_label"].isin(["summary", "min_value", "max_value"])]
        # Drop unwanted columns
        summary_df = summary_df.drop(["full_name", "interviewee_wealth_group_category", "column"], axis="columns")
        # Pivot the value, min_value and max_value back into columns
        summary_df = verbose_pivot(
            summary_df,
            values="value",
            index=[col for col in summary_df.columns if col not in ["interview_label", "value"]],
            columns="interview_label",
        )
        summary_df.columns.name = None
        summary_df = summary_df.rename(columns={"summary": "value"})
        # Add the source column
        summary_df["reference_type"] = WealthGroupCharacteristicValue.CharacteristicReference.SUMMARY
        # Summary Wealth Group Characteristic Values don't have a community
        summary_df["full_name"] = None

        # Drop the summary data from the main dataframe
        char_df = char_df[~char_df["interview_label"].isin(["summary", "min_value", "max_value"])]

        # Add the source column to the main dataframe
        # Rows with an interviewee wealth group category are from the Wealth Group-level Form 4 interview.
        char_df["reference_type"] = char_df["interviewee_wealth_group_category"].where(
            char_df["interviewee_wealth_group_category"] == "",
            WealthGroupCharacteristicValue.CharacteristicReference.WEALTH_GROUP,
        )
        # Rows without an interviewee wealth group category are from the Community-level Form 3 interview.
        char_df["reference_type"] = char_df["reference_type"].mask(
            char_df["reference_type"] == "", WealthGroupCharacteristicValue.CharacteristicReference.COMMUNITY
        )

        # Community and Wealth Group interviews don't have min_value or max_value
        char_df["max_value"] = None
        char_df["min_value"] = None

        # Add the summary df back into the main df, keeping only the columns
        char_df = pd.concat([char_df, summary_df])

        # The percentage of households should be stored as a number between 1 and 100,
        # but may be stored in the BSS (particularly in the summary column) as a
        # decimal fraction between 0 and 1, so correct those values
        char_df.loc[
            char_df["wealth_characteristic"].eq("percentage of households")
            & (pd.to_numeric(char_df["value"], errors="coerce") < 1),
            "value",
        ] *= 100

        # Add the natural key for the Wealth Group
        char_df["wealth_group"] = char_df[["wealth_group_category", "full_name"]].apply(
            lambda row: [
                metadata["code"],
                metadata["reference_year_end_date"].date().isoformat(),
                row["wealth_group_category"],
                row["full_name"],
            ],
            axis="columns",
        )

        # The Wealth Groups also include the household size and percentage of households,
        # so derive them from the Wealth Group Characteristic Values. We want the rows from char_df
        # for those characteristics, where the source is either the Wealth Group Interview
        # or the Summary.
        wealth_group_df = char_df[
            char_df["wealth_characteristic"].isin(["percentage of households", "household size"])
            & char_df["reference_type"].isin(
                [
                    WealthGroupCharacteristicValue.CharacteristicReference.WEALTH_GROUP,
                    WealthGroupCharacteristicValue.CharacteristicReference.SUMMARY,
                ]
            )
        ]

        # Make sure that the names in the Wealth Group-level interviews match
        # the names in the in the Community-level interviews that were used to
        # create the Community records
        unmatched_full_names = wealth_group_df[
            pd.notna(wealth_group_df["full_name"]) & ~wealth_group_df["full_name"].isin(community_df.full_name)
        ][["full_name", "column", "row"]]
        if not unmatched_full_names.empty:
            raise PipelineError(
                "Unmatched Community.full_name in Wealth Group interviews:\n%s" % unmatched_full_names.to_markdown()
            )

        # Drop unwanted columns
        wealth_group_df = wealth_group_df.drop(
            ["wealth_characteristic_original", "product", "product_original"], axis="columns"
        )

        # Pivot the percentage of households and household size back into columns
        wealth_group_df = verbose_pivot(
            wealth_group_df,
            values="value",
            index=["wealth_group_category", "full_name"],
            columns="wealth_characteristic",
        )
        wealth_group_df = wealth_group_df.rename(
            columns={
                "percentage of households": "percentage_of_households",
                "household size": "average_household_size",
            }
        )

        # Make sure that we have a Wealth Group for every combination of Wealth Group Category and Community,
        # because sometimes there is no Wealth Group-level (Form 4) data but there is Community-level (Form 3)
        # data for the Wealth Group.
        # Generate all possible combinations of unique 'full_name' and 'wealth_group_category'
        all_combinations = pd.DataFrame(
            list(
                itertools.product(
                    wealth_group_df["full_name"].unique(), wealth_group_df["wealth_group_category"].unique()
                )
            ),
            columns=["full_name", "wealth_group_category"],
        )
        wealth_group_df = pd.merge(
            all_combinations, wealth_group_df, on=["full_name", "wealth_group_category"], how="outer"
        )

        # Add the natural key for the livelihood zone baseline and community
        wealth_group_df["livelihood_zone_baseline"] = wealth_group_df["full_name"].apply(
            lambda full_name: [metadata["code"], metadata["reference_year_end_date"].date().isoformat()]
        )
        wealth_group_df["community"] = wealth_group_df["full_name"].apply(
            lambda full_name: None
            if pd.isna(full_name)
            else [metadata["code"], metadata["reference_year_end_date"].date().isoformat(), full_name]
        )
        wealth_group_df = wealth_group_df.drop(columns="full_name")

        # Add the Wealth Groups to the result
        result["WealthGroup"] = wealth_group_df.to_dict(orient="records")

        # Add the Wealth Group Characteristic Values to the result
        result["WealthGroupCharacteristicValue"] = char_df.to_dict(orient="records")

        return result


@requires(normalize_wb=NormalizeWB)
class NormalizeBaseline(luigi.Task):
    """
    Normalize a Baseline read from a BSS into a standard format.
    """

    def output(self):
        return IntermediateTarget(path_from_task(self) + ".json", format=JSON, timeout=3600)

    def run(self):
        data = self.input()["normalize_wb"].open().read()

        with self.output().open("w") as output:
            output.write(data)


@requires(NormalizeBaseline)
class ValidateBaseline(luigi.Task):
    """
    Validate a Baseline read from a BSS, and raise an exception if the BSS cannot be imported.
    """

    def output(self):
        return IntermediateTarget(path_from_task(self) + ".json", format=JSON, timeout=3600)

    def run(self):
        data = self.input().open().read()

        # Run the validation
        self.validate(data)

        # Save the validated data
        with self.output().open("w") as output:
            output.write(data)

    def validate(self, data) -> None:
        """
        Validate the normalized data and raise a PipelineError if it isn't valid
        """
        errors = []
        for model_name, instances in data.items():
            model = class_from_name(f"baseline.models.{model_name}")
            # Ignore non-`baseline` models.
            # The normalized data can include other models, e.g. from `metadata`,
            # but those records must be loaded separately to ensure that they are properly reviewed.
            if model._meta.app_label == "baseline":
                df = pd.DataFrame.from_records(instances)

                for field in model._meta.get_fields():
                    # Only validate foreign keys to model that aren't in the baseline app.
                    # The data can contain natural key references to parent models within
                    # baseline, and the fixture will load the parent and the child in the
                    # same transaction in ImportBaseline below.
                    if isinstance(field, ForeignKey) and field.related_model._meta.app_label != "baseline":
                        column = field.name
                        if df[column].isnull().values.any():
                            unmatched_metadata = get_unmatched_metadata(df, column)
                            for value in unmatched_metadata:
                                error = f"Unmatched remote metadata for {model_name} with {column} '{value}'"
                                errors.append(error)

        if errors:
            raise PipelineError(f"{self}: Missing or inconsistent metadata in BSS", errors=errors)


@requires(ValidateBaseline)
class BuildFixture(luigi.Task):
    """
    Convert the validated data into a JSON-format Django fixture.
    """

    def output(self):
        # The file extension must be verbose_json to match the name of the serialization format that we want to use.
        return IntermediateTarget(path_from_task(self) + ".verbose_json", format=JSON, timeout=3600)

    def run(self):
        with self.input().open() as input:
            data = input.read()

        fixture = self.create_fixture(data)

        with self.output().open("w") as output:
            output.write(fixture)

    def create_fixture(self, data):
        """
        Convert the validated data into a JSON-format Django fixture.

        The resulting fixture can be loaded using `loaddata`. The fixture
        uses natural keys to allow it to load all the data in one call
        without needing sequentially insert model instances and keep track
        of the primary keys for each one.
        """
        fixture = []

        for model_name, instances in data.items():
            model = class_from_name(f"baseline.models.{model_name}")
            # Ignore non-`baseline` models.
            # The validated data can include other models, e.g. from `metadata`,
            # but those records must be loaded separately to ensure that they are properly reviewed.
            if model._meta.app_label == "baseline":
                valid_field_names = [field.name for field in model._meta.get_fields()]
                for field_values in instances:
                    record = {
                        "model": str(model._meta),  # returns, e.g; baseline.livelihoodzone
                    }
                    if not hasattr(model, "natural_key"):
                        # This model doesn't use a natural key, so we need to specify the primary key separately
                        try:
                            record["pk"] = field_values.pop(model._meta.pk.name)
                        except KeyError:
                            raise PipelineError(
                                "Model %s doesn't support natural keys, and the data doesn't contain the primary key field '%s'"  # NOQA: E501
                                % (model_name, model._meta.pk.name)
                            )
                    # Discard any fields that aren't in model - they are probably left over from dataframe
                    # manipulation, such as the "_original" fields from metadata lookups, and replace nan with None.
                    record["fields"] = {}
                    for field, value in field_values.items():
                        if field in valid_field_names:
                            try:
                                if pd.isna(value):
                                    value = None
                            except ValueError:
                                # value is a list, probably a natural key, and raises:
                                # "The truth value of an array with more than one element is ambiguous"
                                pass
                            record["fields"][field] = value
                    fixture.append(record)

        return fixture


@requires(BuildFixture)
class ImportBaseline(luigi.Task):
    """
    Import a Baseline, including all child tables from a JSON fixture created from a BSS.
    """

    def output(self):
        return IntermediateTarget(path_from_task(self) + ".json", format=JSON, timeout=3600)

    def run(self):
        buffer = io.StringIO()

        call_command(verbose_load_data.Command(), self.input().path, verbosity=2, format="verbose_json", stdout=buffer)

        with self.output().open("w") as output:
            # Data is a JSON object
            output.write(buffer.getvalue())


class ImportBaselines(luigi.WrapperTask):
    """
    Import multiple Baselines, including all child tables from a JSON fixture created from the BSSs.
    """

    bss_paths = luigi.ListParameter(default=[], description="List of path to the BSS files")
    metadata_path = luigi.Parameter(description="Path to the BSS metadata")
    corrections_path = luigi.Parameter(default="", description="Path to the BSS corrections")

    def requires(self):
        return [
            ImportBaseline(bss_path=bss_path, metadata_path=self.metadata_path, corrections_path=self.corrections_path)
            for bss_path in self.bss_paths
        ]


class ImportAllBaselines(luigi.WrapperTask):
    """
    Import all the Baselines stored as BSS files within a folder, including in subfolfders.
    """

    root_path = luigi.Parameter(description="Path to the root folder containing the BSSs")
    metadata_path = luigi.Parameter(description="Path to the BSS metadata")
    corrections_path = luigi.Parameter(default="", description="Path to the BSS corrections")

    def requires(self):
        root_target = LocalTarget(Path(self.root_path).expanduser().absolute(), format=luigi.format.Nop)
        if not root_target.exists():
            root_target = GoogleDriveTarget(
                self.root_path,
                format=luigi.format.Nop,
                mimetype="application/vnd.google-apps.folder",
            )
            if not root_target.exists():
                raise PipelineError("No local or Google Drive folder matching %s" % self.root_path)
        for path in root_target.fs.listdir(self.root_path):
            yield ImportBaseline(
                bss_path=path, metadata_path=self.metadata_path, corrections_path=self.corrections_path
            )
