#!/usr/bin/env python3
"""
Azure OpenAI Agent PoC with function calling + JSON label cache.

- Tools:
    1) get_cell(sheet, coord)                -> {node_id, is_formula, formula, value}
    2) list_precedents(sheet, coord)         -> list of {node_id, sheet, coord, from_range?}
    3) get_or_ask_label(node_id, value, ...) -> label string (checks JSON first; if absent, prompts human and saves)
- Flow:
    User gives workbook + target (sheet, cell). Model decides how to recurse, asks for labels for static leaves,
    then returns a final *summary* text.

Run:
    python llm_formula_recurse.py data/LIAS_Senegal.xlsx B E277 --labels cell_labels.json --max-range 150
"""

import argparse
import json
import os
import re
import sys
from typing import Dict, List, Optional, Tuple
import openai
import ssl
import certifi
import httpx

import openpyxl
from azure.identity import DefaultAzureCredential, get_bearer_token_provider

# =========================
# Excel helpers (A1 parsing + IO)
# =========================

A1_REF_RE = re.compile(
    r"(?:(?P<sheet>'[^']+'|[A-Za-z0-9_ .-]+)\!)?"
    r"(?P<ref>\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?)"
)

def col_to_num(col: str) -> int:
    col = col.replace("$", "").upper()
    n = 0
    for ch in col:
        if "A" <= ch <= "Z":
            n = n * 26 + (ord(ch) - 64)
    return n

def num_to_col(n: int) -> str:
    s = []
    while n:
        n, r = divmod(n - 1, 26)
        s.append(chr(r + 65))
    return "".join(reversed(s))

def split_a1(cell: str) -> Tuple[str, str]:
    cell = cell.upper()
    i = 0
    while i < len(cell) and (cell[i] == "$" or cell[i].isalpha()):
        i += 1
    return cell[:i], cell[i:]

def expand_range(a: str, b: str) -> List[str]:
    col_a, row_a = split_a1(a)
    col_b, row_b = split_a1(b)
    c1, c2 = col_to_num(col_a), col_to_num(col_b)
    r1, r2 = int(row_a.replace("$", "")), int(row_b.replace("$", ""))
    out = []
    for c in range(min(c1, c2), max(c1, c2) + 1):
        for r in range(min(r1, r2), max(r1, r2) + 1):
            out.append(f"{num_to_col(c)}{r}")
    return out

def parse_formula_refs(formula: str, current_sheet: Optional[str]) -> List[Dict[str, str]]:
    items: List[Dict[str, str]] = []
    if not formula:
        return items
    for m in A1_REF_RE.finditer(formula):
        sheet = m.group("sheet")
        ref = m.group("ref").upper()
        if sheet:
            sheet = sheet.strip("'")
        else:
            sheet = current_sheet
        if ":" in ref:
            a, b = ref.split(":")
            items.append({"type":"range","sheet":sheet,"ref":ref,"a":a,"b":b})
        else:
            items.append({"type":"cell","sheet":sheet,"ref":ref})
    return items

def open_wb_formulas(path: str):
    return openpyxl.load_workbook(path, data_only=False, read_only=True)

def open_wb_values(path: str):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)

def _strip_abs(a1: str) -> str:
    return a1.replace("$", "").upper()

def get_formula_text(ws, coord: str) -> Optional[str]:
    c = ws[_strip_abs(coord)]
    if c is None:
        return None
    val = c.value
    if val is None:
        return None
    s = str(val)
    if getattr(c, "data_type", None) == "f" or s.startswith("="):
        return s[1:] if s.startswith("=") else s
    return None

def get_formula_with_eq(ws, coord: str) -> Optional[str]:
    f = get_formula_text(ws, coord)
    if f is None:
        return None
    return f if f.startswith("=") else "=" + f

def get_cell_value(ws, coord: str):
    c = ws[_strip_abs(coord)]
    return None if c is None else c.value

# =========================
# Tool implementations (backed by Excel)
# =========================

class ExcelAgentTools:
    def __init__(self, workbook_path: str, labels_path: str, max_range: int = 150):
        self.path = workbook_path
        self.labels_path = labels_path
        self.max_range = max_range
        self._wf = open_wb_formulas(workbook_path)
        self._wv = open_wb_values(workbook_path)
        self.labels: Dict[str, str] = self._load_labels()

    # ---------- labels ----------
    def _load_labels(self) -> Dict[str, str]:
        if not self.labels_path:
            return {}
        try:
            with open(self.labels_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except FileNotFoundError:
            return {}
        except Exception as e:
            print(f"[warn] Could not read labels JSON: {e}", file=sys.stderr)
            return {}

    def _save_labels(self):
        if not self.labels_path:
            return
        try:
            os.makedirs(os.path.dirname(self.labels_path) or ".", exist_ok=True)
            with open(self.labels_path, "w", encoding="utf-8") as f:
                json.dump(self.labels, f, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"[warn] Could not save labels JSON: {e}", file=sys.stderr)

    # ---------- tools ----------
    def get_cell(self, sheet: str, coord: str) -> Dict:
        sheet = sheet
        coord = coord.upper()
        if sheet not in self._wf.sheetnames:
            return {"error": f"Sheet '{sheet}' not found"}
        ws_f = self._wf[sheet]
        ws_v = self._wv[sheet]
        f = get_formula_text(ws_f, coord)
        feq = get_formula_with_eq(ws_f, coord)
        val = get_cell_value(ws_v, coord)
        return {
            "node_id": f"{sheet}!{coord}",
            "sheet": sheet,
            "coord": coord,
            "is_formula": f is not None,
            "formula": f,
            "display_formula": feq,
            "value": val
        }

    def list_precedents(self, sheet: str, coord: str) -> Dict:
        """Expand refs; limit big ranges to first max_range cells."""
        if sheet not in self._wf.sheetnames:
            return {"error": f"Sheet '{sheet}' not found"}
        ws_f = self._wf[sheet]
        f = get_formula_text(ws_f, coord)
        if f is None:
            return {"precedents": []}  # static
        refs = parse_formula_refs(f, current_sheet=sheet)
        out = []
        for it in refs:
            if it["type"] == "cell":
                out.append({
                    "node_id": f"{it['sheet']}!{it['ref']}",
                    "sheet": it["sheet"],
                    "coord": it["ref"],
                })
            else:
                cells = expand_range(it["a"], it["b"])
                capped = False
                if len(cells) > self.max_range:
                    capped = True
                    cells = cells[:self.max_range]
                for c in cells:
                    out.append({
                        "node_id": f"{it['sheet']}!{c}",
                        "sheet": it["sheet"],
                        "coord": c,
                        "from_range": f"{it['sheet']}!{it['ref']}",
                        "range_capped": capped
                    })
        return {"precedents": out}

    def get_or_ask_label(self, node_id: str, value) -> Dict:
        """Return existing label if present, otherwise prompt user and save."""
        if node_id in self.labels and self.labels[node_id]:
            return {"node_id": node_id, "label": self.labels[node_id], "source": "cache"}
        # Prompt human
        val_str = "(empty)" if value is None else repr(value)
        print(f"\nLabel needed for {node_id} (value={val_str})")
        label = input("  What does this value represent? (short phrase): ").strip()
        self.labels[node_id] = label
        self._save_labels()
        return {"node_id": node_id, "label": label, "source": "user"}

    def close(self):
        try: self._wf.close()
        except: pass
        try: self._wv.close()
        except: pass

# =========================
# Azure OpenAI client
# =========================

def initialize_azure_client(division="ts", region="eastus2", api_version="2024-10-21"):
    openai_endpoints = {
            'ts': {
                'eastus':'https://air-ts-eastus.openai.azure.com/',
                'eastus2':'https://air-ts-eastus2.openai.azure.com/',
                'northcentralus':'https://air-poc-northcentralus.openai.azure.com/',
            },
            'ps': {
                'eastus':'https://air-ps-eastus.openai.azure.com/',
                'eastus2':'https://air-ps-eastus2.openai.azure.com/',
                'northcentralus':'https://air-poc-northcentralus.openai.azure.com/'
            },
        }
    openai_endpoint = openai_endpoints[division][region]
    token_provider = get_bearer_token_provider(DefaultAzureCredential(), "https://cognitiveservices.azure.com/.default")
    ctx = ssl.create_default_context(cafile=os.environ.get('REQUESTS_CA_BUNDLE', certifi.where()))
    httpx_client = httpx.Client(verify=ctx)
    openai_client = openai.AzureOpenAI(
        api_version=api_version,
        azure_endpoint=openai_endpoint,
        azure_ad_token_provider=token_provider,
        http_client=httpx_client
    )
    return openai_client

# =========================
# Tool specs (for function calling)
# =========================

TOOLS = [
    {
        "type": "function",
        "function": {
            "name": "get_cell",
            "description": "Get info for a cell (is it a formula? what's its formula/value?).",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "coord": {"type": "string"}
                },
                "required": ["sheet", "coord"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "list_precedents",
            "description": "List precedent cells for a formula (expand ranges up to a limit).",
            "parameters": {
                "type": "object",
                "properties": {
                    "sheet": {"type": "string"},
                    "coord": {"type": "string"}
                },
                "required": ["sheet", "coord"]
            }
        }
    },
    {
        "type": "function",
        "function": {
            "name": "get_or_ask_label",
            "description": "Return a semantic label for a static value. Use cache if present; otherwise ask the user and save.",
            "parameters": {
                "type": "object",
                "properties": {
                    "node_id": {"type": "string"},
                    "value": {}
                },
                "required": ["node_id", "value"]
            }
        }
    }
]

SYSTEM_PROMPT = """You are an Excel tracing agent. The workbook you'll be working with is a Livelihoods and Impact Analysis Spreadsheet (LIAS).
Goal:
- Given a target Sheet!Cell, recursively trace precedents until only static (non-formula) values remain.
- For each static value, call get_or_ask_label to retrieve a human-provided semantic meaning (cached if known).
- When you have labels for all static inputs referenced (directly or transitively) by the target, produce a clear natural-language summary describing how the target is computed from those inputs.

Ground rules:
- ALWAYS use get_cell to check if a node is a formula before assuming anything.
- For formula cells, use list_precedents to step upstream.
- For static cells, use get_or_ask_label.
- Keep calling tools until all static leaves are labeled; then return a final explanation (no further tool calls).
- The final message should be a clear, readable explanation in plain English, interpreted in the context of livelihood analysis. Do not include JSON in the final message.
"""

# =========================
# Controller loop
# =========================

def run_agent(workbook: str, sheet: str, cell: str, context: str, labels_path: str, max_range: int, max_rounds: int = 30):

    client = initialize_azure_client()
    tools = ExcelAgentTools(workbook, labels_path, max_range=max_range)

    try:
        messages = [
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": f"Workbook path: {workbook}\nTarget: {sheet}!{cell.upper()}.\nContext of target cell: {context}"},
        ]

        for round_idx in range(max_rounds):
            resp = client.chat.completions.create(
                model="gpt-4o",  # Azure: use *deployment name*
                messages=messages,
                tools=TOOLS,
                tool_choice="auto",
                temperature=0.1,
            )
            msg = resp.choices[0].message

            if not msg.tool_calls:
                print("\n=== SUMMARY ===\n")
                print(msg.content.strip())
                return

            # Handle tool calls (may be multiple)
            messages.append({"role": "assistant", "content": msg.content or "", "tool_calls": msg.tool_calls})

            for tc in msg.tool_calls:
                name = tc.function.name
                args_json = tc.function.arguments or "{}"
                try:
                    params = json.loads(args_json)
                except Exception:
                    params = {}

                # Route to our Python tool implementations
                if name == "get_cell":
                    result = tools.get_cell(params.get("sheet", ""), params.get("coord", ""))
                elif name == "list_precedents":
                    result = tools.list_precedents(params.get("sheet", ""), params.get("coord", ""))
                elif name == "get_or_ask_label":
                    result = tools.get_or_ask_label(params.get("node_id", ""), params.get("value", None))
                else:
                    result = {"error": f"Unknown tool: {name}"}

                # Append tool result
                messages.append({
                    "role": "tool",
                    "tool_call_id": tc.id,
                    "name": name,
                    "content": json.dumps(result, ensure_ascii=False)
                })

        print("[warn] Max rounds reached without final summary.")
    finally:
        tools.close()


def main():
    ap = argparse.ArgumentParser(description="Azure OpenAI Excel tracing agent (function-calling PoC)")
    ap.add_argument("workbook", help="Path to .xlsx")
    ap.add_argument("sheet", help="Sheet name, e.g. S")
    ap.add_argument("cell", help="Cell, e.g. P23 or $P$23")
    ap.add_argument("context", help="A description of what the cell represents")
    ap.add_argument("--labels", default="labels.json", help="Path to JSON label store")
    ap.add_argument("--max-range", type=int, default=150, help="Max range expansion per reference")
    args = ap.parse_args()

    run_agent(args.workbook, args.sheet, args.cell, args.context, labels_path=args.labels, max_range=args.max_range)

if __name__ == "__main__":
    main()
