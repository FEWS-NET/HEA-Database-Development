from typing import List, Optional
import openpyxl

def open_workbook(path: str, *, read_only: bool = False):
    return openpyxl.load_workbook(path, data_only=False, read_only=read_only)

# NEW: open a values-view workbook (uses cached values saved by Excel)
def open_workbook_values(path: str):
    return openpyxl.load_workbook(path, data_only=True, read_only=True)

def list_sheets(path: str) -> List[str]:
    wb = open_workbook(path, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()

def _strip_abs(a1: str) -> str:
    return a1.replace("$", "").upper()

def get_formula_text(ws, coord: str) -> Optional[str]:
    c = ws[_strip_abs(coord)]
    if c is None:
        return None
    val = c.value
    if val is None:
        return None
    s = str(val)
    if getattr(c, "data_type", None) == "f" or s.startswith("="):
        return s[1:] if s.startswith("=") else s
    return None

def get_formula_with_equals(ws, coord: str) -> Optional[str]:
    f = get_formula_text(ws, coord)
    if f is None:
        return None
    return f if f.startswith("=") else "=" + f

# NEW: read the cached value (evaluated in Excel, if available)
def get_cell_value(ws, coord: str):
    c = ws[_strip_abs(coord)]
    return None if c is None else c.value
