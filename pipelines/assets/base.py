"""
Base functions for performing operations on BSS spreadsheets.
"""

import numbers
import os
from io import BytesIO

import django
import msoffcrypto
import openpyxl
import pandas as pd
import xlrd
import xlwt
from dagster import (
    AssetExecutionContext,
    DagsterEventType,
    EventRecordsFilter,
    MetadataValue,
    Output,
    asset,
)
from django.db.models import F
from gdrivefs.core import GoogleDriveFile
from googletrans import Translator
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, rows_from_range
from upath import UPath
from xlutils.copy import copy as copy_xls

from ..configs import BSSMetadataConfig
from ..partitions import bss_files_partitions_def, bss_instances_partitions_def
from ..utils import get_index, prepare_lookup

# set the default Django settings module
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hea.settings.production")

# Configure Django with our custom settings before importing any Django classes
django.setup()

from baseline.models import LivelihoodZoneBaseline  # NOQA: E402
from common.lookups import CountryLookup  # NOQA: E402
from metadata.models import ActivityLabel, WealthCharacteristicLabel  # NOQA: E402

# Map of cell values in Column A to the language of the BSS.
# For WB, Data, Data2 and Data3 worksheets these appear in row 3
# For 'Exp factors' they appear in row 2
# For 'Seas Cal' they appear in row 1 or 3.
LANGS = {
    "wealth group": "en",
    "group de richesse": "fr",
    "groupe de richesse": "fr",
    "groupe socio-economique": "fr",
    "group socio-economique": "fr",
    # 'Exp factors' (row 2)
    "expandability parameters": "en",
    "paramètres d'extensibilité (expandability factors)": "fr",
    "parametres d'utilisation maximale (expandability)": "fr",
    # 'Seas Cal' (row 1)
    "seasonal calendar": "en",
    "calendrier saisonnier": "fr",
    "calendrier saisonier": "fr",  # typo variant found in some BSSs
    # 'Seas Cal' (row 3) - French BSSs that omit the title from  row 1
    "activités/évenements": "fr",
    "activites/evenements": "fr",  # no-accents variant
}

# List of labels that indicate the start of the summary columns from row 3 in the Data, Data, and Data3 worksheets
SUMMARY_LABELS = [
    "BASELINE",
    "BASE DE RÉFÉRENCE",
    "REFERENCE",
    "REFERENCE DE BASE",
    "BASE DE RÉFÉRENAE",
    "range",
    "interval",
    "intervales",  # 2023 Mali BSSs
    # Seas Cal worksheets
    "Results",
    "Synthèse",
]


@asset
def bss_metadata(context: AssetExecutionContext, config: BSSMetadataConfig) -> Output[pd.DataFrame]:
    """
    A DataFrame containing the BSS Metadata.
    """
    p = UPath(config.bss_metadata_workbook, **config.bss_metadata_storage_options)

    with p.fs.open(p.path, mode="rb", cache_type="bytes") as f:
        # Google Sheets have to exported rather than read directly
        if isinstance(f, GoogleDriveFile) and (f.details["mimeType"] == "application/vnd.google-apps.spreadsheet"):
            f = BytesIO(p.fs.export(p.path, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))

        df = pd.read_excel(f, sheet_name="Metadata", engine="openpyxl")

    df["reference_year_start_date"] = df["reference_year_start_date"].dt.date
    df["reference_year_end_date"] = df["reference_year_end_date"].dt.date
    df["valid_from_date"] = df["valid_from_date"].dt.date
    df["valid_to_date"] = df["valid_to_date"].dt.date
    df["data_collection_start_date"] = df["data_collection_start_date"].dt.date
    df["data_collection_end_date"] = df["data_collection_end_date"].dt.date
    df["publication_date"] = df["publication_date"].dt.date

    # Add the partition key and a flag indicating whether the BSS file exists
    partition_key_df = df[["country_id", "code", "reference_year_end_date"]]
    partition_key_df = CountryLookup().get_instances(partition_key_df, "country_id")
    df["partition_key"] = partition_key_df[["country_id", "code", "reference_year_end_date"]].apply(
        lambda x: f"{x.iloc[0].iso_en_ro_proper}~{x.iloc[1]}~{x.iloc[2]}", axis="columns"
    )

    partitions = bss_files_partitions_def.get_partition_keys(dynamic_partitions_store=context.instance)

    partitions_to_delete = [partition for partition in partitions if partition not in df["partition_key"].tolist()]
    for partition_key in partitions_to_delete:
        context.instance.delete_dynamic_partition(bss_files_partitions_def.name, partition_key)

    new_partitions = df[df["bss_exists"] & ~df["partition_key"].isin(partitions)]["partition_key"].tolist()
    context.instance.add_dynamic_partitions(bss_files_partitions_def.name, new_partitions)

    return Output(
        df,
        metadata={
            "num_baselines": len(df),
            "preview": MetadataValue.md(
                df[["partition_key", "bss_path", "status", "name_en", "main_livelihood_category_id"]]
                .head(config.preview_rows)
                .to_markdown(index=False)
                .replace("~", "\\~")  # Escape the ~ in the partition_key, otherwise it is rendered as strikethrough
            ),
        },
    )


@asset
def completed_bss_metadata(config: BSSMetadataConfig, bss_metadata) -> Output[pd.DataFrame]:
    """
    A DataFrame containing the BSS Metadata that has been completed sufficiently to allow the BSS to be loaded.
    """
    required_columns = [
        "bss_path",
        "code",
        "country_id",
        "source_organization",
        "name_en",
        "main_livelihood_category_id",
        "currency_id",
        "reference_year_start_date",
        "reference_year_end_date",
        "valid_from_date",
    ]
    bss_metadata = bss_metadata[bss_metadata["bss_exists"]].sort_values(by="bss_path")
    mask = bss_metadata[required_columns].map(lambda x: x == "")

    # Drop rows where any of the specified columns have empty strings
    complete_df = bss_metadata[~mask.any(axis="columns")]
    incomplete_df = bss_metadata[mask.any(axis="columns")]

    return Output(
        complete_df,
        metadata={
            "num_baselines": len(bss_metadata),
            "num_complete": len(complete_df),
            "num_incomplete": len(incomplete_df),
            "complete": MetadataValue.md(
                complete_df[["bss_path", "name_en", "main_livelihood_category_id"]].to_markdown()
            ),
            "incomplete": MetadataValue.md(
                incomplete_df[["bss_path", "status", "name_en", "main_livelihood_category_id"]].to_markdown()
            ),
            "preview": MetadataValue.md(
                complete_df[["bss_path", "status", "name_en", "main_livelihood_category_id"]]
                .head(config.preview_rows)
                .to_markdown()
            ),
        },
    )


@asset(partitions_def=bss_files_partitions_def)
def original_files(context: AssetExecutionContext, config: BSSMetadataConfig, bss_metadata) -> Output[BytesIO]:
    """
    Original BSS files as used by the HEA community.
    """
    partition_key = context.asset_partition_key_for_output()
    partition_metadata = bss_metadata[bss_metadata["partition_key"] == partition_key]
    if partition_metadata.empty:
        raise ValueError(f"No BSS metadata found for partition key `{partition_key}`")
    bss_path = partition_metadata["bss_path"].iloc[0]
    p = UPath(config.bss_files_folder, bss_path, **config.bss_files_storage_options)

    # Prepare the metadata for the output
    output_metadata = {"bss_path": p.path}

    with p.fs.open(p.path, mode="rb") as f:
        return Output(BytesIO(f.read()), metadata=output_metadata)


@asset(partitions_def=bss_instances_partitions_def)
def corrected_files(context: AssetExecutionContext, config: BSSMetadataConfig) -> Output[BytesIO]:
    """
    BSS files with any necessary corrections applied.
    """
    partition_key = context.asset_partition_key_for_output()
    livelihood_zone_baseline = LivelihoodZoneBaseline.objects.get_by_natural_key(*partition_key.split("~")[1:])

    def validate_previous_value(worksheet_name, cell_reference, expected_previous_value, previous_value):
        """
        Inline function to validate the existing value of a cell is the expected one, prior to correcting it.
        """
        # "#N/A" is inconsistently loaded as nan, even when copied and pasted in Excel or GSheets
        if not isinstance(previous_value, numbers.Number):
            previous_value = str(previous_value).strip()
            expected_previous_value = str(expected_previous_value).strip()
            if previous_value in ["None", "nan", "#N/A", "N/A"]:
                previous_value = ""
            if expected_previous_value in ["None", "nan", "#N/A", "N/A"]:
                expected_previous_value = ""
        else:
            # Provide similar precisions between the Google sheet and original Excel
            expected_previous_value = round(expected_previous_value, 6)
            previous_value = round(previous_value, 6)

        if expected_previous_value != previous_value:
            raise ValueError(
                "Unexpected prior value in source BSS. "
                f"BSS `{partition_key}`, cell `'{worksheet_name}'!{cell_reference}`, "
                f"value found `{previous_value}`, expected `{expected_previous_value}`."
            )

    # Find the corrections for this BSS
    corrections_df = pd.DataFrame.from_records(
        livelihood_zone_baseline.corrections.all()
        .order_by("worksheet_name", "cell_range")
        .values(
            "worksheet_name", "cell_range", "previous_value", "value", "author__username", "correction_date", "comment"
        )
    )

    # Prepare the metadata for the output
    output_metadata = {
        "num_corrections": len(corrections_df),
        "corrections": MetadataValue.md(corrections_df.to_markdown(index=False)),
    }

    # Open the file to force it to be saved to the local media directory in case it only exists in the database.
    with livelihood_zone_baseline.bss.open() as original_file:
        # Excel silently encrypts .xls files if there are locked worksheets, which
        # prevents xlrd from reading them. It uses a hard-coded password to do it!
        # See https://stackoverflow.com/questions/22789951/xlrd-error-workbook-is-encrypted-python-3-2-3
        try:
            wb_msoffcrypto_file = None
            xlrd_wb = xlrd.open_workbook(file_contents=original_file.read(), formatting_info=True, on_demand=True)
            # xlrd can only read XLS files, so we need to use xlutils.copy_xls to create something we can edit
            wb = copy_xls(xlrd_wb)
        except xlrd.biffh.XLRDError as e:
            if str(e) == "Workbook is encrypted":
                # Try and unencrypt workbook with magic password
                wb_msoffcrypto_file = msoffcrypto.OfficeFile(original_file)
                try:
                    # Yes, this is actually a thing!
                    wb_msoffcrypto_file.load_key(password="VelvetSweatshop")
                except Exception as e:
                    raise RuntimeError("Cannot process encrypted workbook") from e
                # Magic Excel password worked
                buffer = BytesIO()
                wb_msoffcrypto_file.decrypt(buffer)
                xlrd_wb = xlrd.open_workbook(file_contents=buffer.getvalue(), formatting_info=True, on_demand=True)
                # xlrd can only read XLS files, so we need to use xlutils.copy_xls to create something we can edit
                wb = copy_xls(xlrd_wb)
            else:
                # We assume that the file is an xlsx file that can be read by openpyxl
                xlrd_wb = None  # Required to suppress spurious unbound variable errors from Pyright
                # Need data_only=True to get the values of cells that contain formulas
                wb = openpyxl.load_workbook(original_file, data_only=True)

        if corrections_df.empty and not wb_msoffcrypto_file:
            # No corrections and file not encrypted, so just leave the file unaltered
            return Output(BytesIO(original_file.open().read()), metadata=output_metadata)

    # Apply the corrections
    for correction in corrections_df.itertuples():
        try:
            for row in rows_from_range(correction.cell_range):
                for cell in row:
                    if isinstance(wb, xlwt.Workbook):
                        row, col = coordinate_to_tuple(cell)
                        previous_value = xlrd_wb.sheet_by_name(correction.worksheet_name).cell_value(row - 1, col - 1)
                        if (
                            xlrd_wb.sheet_by_name(correction.worksheet_name).cell(row - 1, col - 1).ctype
                            == xlrd.XL_CELL_ERROR
                        ):
                            # xlrd.error_text_from_code returns, eg, "#N/A"
                            previous_value = xlrd.error_text_from_code[
                                xlrd_wb.sheet_by_name(correction.worksheet_name).cell_value(row - 1, col - 1)
                            ]
                        validate_previous_value(
                            correction.worksheet_name, cell, correction.previous_value, previous_value
                        )
                        # xlwt uses 0-based indexes, but coordinate_to_tuple uses 1-based, so offset the values
                        wb.get_sheet(correction.worksheet_name).write(row - 1, col - 1, correction.value)
                    else:
                        cell = wb[correction.worksheet_name][cell]
                        validate_previous_value(
                            correction.worksheet_name, cell.coordinate, correction.previous_value, cell.value
                        )
                        cell.value = correction.value
                        cell.comment = Comment(
                            f"{correction.author__username} on {correction.correction_date.date().isoformat()}: {correction.comment}",  # NOQA: E501
                            author=correction.author__username,
                        )
        except Exception as e:
            raise ValueError(
                f"Error applying correction to BSS `{partition_key}`, worksheet `{correction.worksheet_name}`, "
                f"range `{correction.cell_range}`"
            ) from e

    buffer = BytesIO()
    wb.save(buffer)
    return Output(buffer, metadata=output_metadata)


def get_bss_dataframe(
    config: BSSMetadataConfig,
    filepath_or_buffer,
    bss_sheet: str,
    start_strings: list[str],
    end_strings: list[str] | None = None,
    header_rows: list[int] | None = [3, 4, 5],  # List of row indexes that contain the Wealth Group and other headers
    end_col: str | None = None,
    num_summary_cols: int | None = None,
    fill_blank_rows: bool = True,
) -> Output[pd.DataFrame]:
    """
    Retrieve a worksheet from a BSS and return it as a DataFrame.

    Uses Excel row numbers, starting from 1, and column letters, starting from A.
    """
    try:
        df = pd.read_excel(filepath_or_buffer, bss_sheet, header=None)
    except ValueError:
        # The requested worksheet does not exist in the file
        return Output(
            pd.DataFrame(),
            metadata={
                "worksheet": bss_sheet,
                "row_count": 0,
                "message": "Worksheet not present in file",
            },
        )

    # Use a 1-based index to match the Excel Row Number
    df.index += 1
    # Set the column names to match Excel
    df.columns = [get_column_letter(col + 1) for col in df.columns]

    if end_col is None:
        # Find the last column before the summary column, which is in row 3
        end_col = get_index(SUMMARY_LABELS, df.loc[3], offset=-1)
    if not end_col:
        raise ValueError(f'No cell containing any of the summary strings: {", ".join(SUMMARY_LABELS)}')

    if num_summary_cols is None:
        # If the number of summary columns wasn't specified, then assume that
        # there is one summary column for each wealth category, from row 3.
        num_summary_cols = df.loc[3, "B":end_col].dropna().nunique()
    end_col = df.columns[df.columns.get_loc(end_col) + num_summary_cols]

    # Find the row index of the start of dataframe by searching for one of the start strings in column A.
    # Some dataframes (e.g. 'Seas Cal' in some French BSSs) place the village/site label in col B rather than col A, so fall
    # back to the columns listed in start_col_fallbacks when the col A search yields nothing.
    start_row = get_index(start_strings, df.loc[1:, "A"])
    if not start_row:
        start_row = get_index(start_strings, df.loc[1:, "B"])
    if not start_row:
        raise ValueError(f"No cell in Column(s) A:B containing any of the start strings: {', '.join(start_strings)}")

    if end_strings:
        # Find the row before the first row that contains an end string
        end_row = get_index(
            end_strings,
            df.loc[start_row:, "A"],
            offset=-1,
        )
        if not end_row:
            # The WB worksheet typically doesn't have a totals section at the bottom, so if we didn't find
            # an end string, then use the last row rather than raising an error.
            end_row = df.index[-1]
    else:
        # Use the last row. We don't use the last row that contains a label, because the WB worksheet contains data
        # for rows that rely on copying down the label in column A from a previous row.
        end_row = df.index[-1]

    # Find the language using the cells in Column A.
    # For WB, Data, Data2 and Data3 worksheets these appear in row 3
    # For 'Exp factors' they appear in row 2
    # For 'Seas Cal' they appear in row 1 or 3.
    lang = None
    for row in [3, 2, 1]:
        try:
            lang = LANGS[df.loc[row, "A"].strip().lower()]
        except (AttributeError, KeyError):
            # There isn't a text label in the cell, or the label isn't in LANGS
            pass
    if not lang:
        raise ValueError(f'No language could be identified from the labels {df.loc[1:3, "A"].tolist()} in Column A')

    # Filter to just the header rows (typically the Wealth Groups) and the main data (e.g. Livelihood Activities)
    header_rows = [row for row in header_rows if row < start_row]
    df = pd.concat(
        [df.loc[header_rows, :end_col] if header_rows else pd.DataFrame(), df.loc[start_row:end_row, :end_col]]
    )

    # Copy the label from the previous cell for rows that have data but have a blank label.
    # For example, sometimes the wealth characteristic label is only filled in for the first wealth category:
    #   |     | A                                              | B                   | C                  |
    #   |----:|:-----------------------------------------------|:--------------------|:-------------------|
    #   |   1 | MALAWI HEA BASELINES 2015                      | Southern Lakeshore  | Southern Lakeshore |
    #   |  18 | Land area owned (acres)                        | VP                  | 1.4                |
    #   |  19 |                                                | P                   | 1.4                |
    #   |  20 |                                                | M                   | 1.4                |
    #   |  21 |                                                | B/O                 | 1.4                |
    #   |  22 | Camels: total owned at start of year           | VP                  | 0                  |
    if fill_blank_rows:
        # We do this by setting the missing values to pd.NA and then using .ffill()
        # Note that we need to replace the None with something else before the mask() and ffill() so that only
        # the masked values are replaced.
        df["A"] = (
            df["A"]
            .replace({None: ""})
            .mask(
                df["A"].isna() & df.loc[:, "B":].notnull().any(axis="columns"),
                pd.NA,
            )
            .ffill()
        )

    # Replace NaN with "" ready for Django
    df = df.fillna("")

    # Create a sample of rows that contain data, because the first rows may not contain any values.
    # For example the Data sheet contains data for Camel's Milk first, which isn't a common Livelihood Activity.
    sample_df = df[df.loc[:, "B":].apply(lambda row: sum((row != 0) & (row != "")), axis="columns") > 0]
    sample_rows = min(len(sample_df), config.preview_rows)

    return Output(
        df,
        metadata={
            "worksheet": bss_sheet,
            "lang": lang,
            "row_count": len(df),
            "datapoint_count": int(
                df.loc[:, "B":].apply(lambda row: sum((row != 0) & (row != "")), axis="columns").sum()
            ),
            "preview": MetadataValue.md(df.head(config.preview_rows).to_markdown()),
            "sample": MetadataValue.md(sample_df.sample(sample_rows).sort_index().to_markdown()),
        },
    )


def get_bss_label_dataframe(
    context: AssetExecutionContext,
    config: BSSMetadataConfig,
    df: pd.DataFrame,
    asset_key: str,
    num_header_rows: int,
    num_label_cols: int = 1,
    num_summary_cols: int | None = None,
) -> Output[pd.DataFrame]:
    """
    Dataframe of Label References for a worksheet in a BSS.
    """
    if df.empty:
        return Output(
            pd.DataFrame(),
            metadata={
                "row_count": "Worksheet not present in file",
            },
        )

    first_community_col = df.columns[num_label_cols]
    if num_summary_cols:
        last_community_col = df.columns[-num_summary_cols - 1]
    else:
        # The summary columns won't have a community, so find the column after the last column with a community name
        last_community_col = df.loc[5][df.loc[5].replace("", pd.NA).notna()].index[-1]

    # Create a numeric representation of the summary columns that we can use to identify rows with summary data.
    first_summary_col = df.columns[df.columns.get_loc(last_community_col) + 1]
    summary_cols = df.loc[:, first_summary_col:].fillna(0).map(lambda x: 0 if isinstance(x, str) else x).astype(float)

    df = df.iloc[num_header_rows:]  # Ignore the header rows
    instance = context.instance
    dataframe_materialization = instance.get_event_records(
        event_records_filter=EventRecordsFilter(
            event_type=DagsterEventType.ASSET_MATERIALIZATION,
            asset_key=context.asset_key_for_input(asset_key),
            asset_partitions=[context.asset_partition_key_for_input(asset_key)],
        ),
        limit=1,
    )[0].asset_materialization

    # Most worksheets have a single label column (i.e. Column A), but the 'Seas Cal' worksheet may have labels in
    # Columns A and B, add the extra labels from Column B as separate rows in the label_df.
    label_df = pd.DataFrame()
    for label_col in df.columns[0:num_label_cols]:
        column_df = pd.DataFrame()
        column_df["label"] = df[label_col]
        column_df["label_lower"] = prepare_lookup(column_df["label"])
        column_df["bss"] = context.asset_partition_key_for_output()
        column_df["lang"] = dataframe_materialization.metadata["lang"].text
        column_df["worksheet"] = dataframe_materialization.metadata["worksheet"].text
        column_df["row_number"] = df.index
        column_df["datapoint_count"] = df.loc[:, first_community_col:].apply(
            lambda row: sum((row != 0) & (row != "")), axis="columns"
        )
        # Create a flag for whether the label/row contains summary data.
        column_df["in_summary"] = summary_cols.sum(axis="columns") > 0
        label_df = pd.concat([label_df, column_df], ignore_index=True)

    # Drop rows where the label is blank and there are no datapoints.
    label_df = label_df[~((label_df["label"] == "") & (label_df["datapoint_count"] == 0))]

    # Drop rows where the label is blank but there is another label for the same row that isn't blank.
    label_df = label_df[
        ~((label_df["label"] == "") & (label_df.groupby("row_number")["label"].transform(lambda x: (x != "").any())))
    ]

    # Avoid double-counting by setting the datapoint_count and in_summary to NA if there is another label for the
    # same row that isn't blank, even if the current row's label is also not blank - because sometimes the 'Seas Cal'
    # worksheet has labels in both Column A and B, and these would otherwise cause the row to be double-counted.
    for col in ["datapoint_count", "in_summary"]:
        label_df[col] = label_df.apply(
            lambda row: (
                pd.NA
                if (
                    label_df.loc[: row.name - 1][
                        (label_df.loc[: row.name - 1]["row_number"] == row["row_number"])
                        & (label_df.loc[: row.name - 1]["label"] != "")
                    ].shape[0]
                    > 0
                )
                else row[col]
            ),
            axis="columns",
        )

    # Sort by row_number to ensure the labels are in the same order as the original worksheet, using a stable sort
    # so that the labels are orderd by row and then column.
    label_df = label_df.sort_values(by="row_number", kind="stable").reset_index(drop=True)

    # Create a sample of rows that contain summary data, because the first rows may not contain any values.
    # For example the Data sheet contains data for Camel's Milk first, which isn't a common Livelihood Activity.
    sample_df = label_df[label_df["in_summary"].fillna(False)]
    sample_rows = min(len(sample_df), config.preview_rows)
    sample_df = sample_df.sample(sample_rows)
    # Filter by row_number so that if the dataframe has more than one label column (e.g. 'Seas Cal') then
    # the sample contains all the labels from the sampled rows.
    sample_df = label_df[label_df["row_number"].isin(sample_df["row_number"])].sort_values(by=["row_number"])

    return Output(
        label_df,
        metadata={
            "num_labels": len(label_df),
            "num_datapoints": int(label_df["datapoint_count"].fillna(0).sum()),
            "num_summaries": int(label_df["in_summary"].fillna(0).astype(int).sum()),
            # Escape the ~ in the partition_key, otherwise it is rendered as strikethrough
            "preview": MetadataValue.md(label_df.head(config.preview_rows).to_markdown().replace("~", "\\~")),
            "sample": MetadataValue.md(sample_df.to_markdown().replace("~", "\\~")),
        },
    )


def get_all_bss_labels_dataframe(
    config: BSSMetadataConfig, label_dataframe: dict[str, pd.DataFrame]
) -> Output[pd.DataFrame]:
    """
    Combined dataframe of the activity labels in use across all BSSs.
    """
    df = pd.concat(list(label_dataframe.values()))

    # Create a sample of rows that contain summary data, because the first rows may not contain any values.
    # For example the Data sheet contains data for Camel's Milk first, which isn't a common Livelihood Activity.
    sample_df = df[df["in_summary"].fillna(False)]
    sample_rows = min(len(sample_df), config.preview_rows)
    sample_df = sample_df.sample(sample_rows)
    # Include any additional labels from the sampled rows by filtering by bss and row_number, so that if the dataframe
    # has more than one label column (e.g. 'Seas Cal') then the sample contains all the labels from the sampled rows.
    bss_row_pairs = set(zip(sample_df["bss"], sample_df["row_number"]))
    sample_df = df[df.apply(lambda row: (row["bss"], row["row_number"]) in bss_row_pairs, axis=1)].sort_values(
        by=["bss", "row_number"]
    )

    return Output(
        df,
        metadata={
            "num_labels": len(df),
            "num_datapoints": int(df["datapoint_count"].sum()),
            "num_summaries": int(df["in_summary"].sum()),
            # Escape the ~ in the partition_key, otherwise it is rendered as strikethrough
            "preview": MetadataValue.md(df.head(config.preview_rows).to_markdown().replace("~", "\\~")),
            "sample": MetadataValue.md(sample_df.to_markdown().replace("~", "\\~")),
        },
    )


def get_summary_bss_label_dataframe(
    config: BSSMetadataConfig, all_labels_dataframe: pd.DataFrame, label_type: str
) -> Output[pd.DataFrame]:
    df = all_labels_dataframe.sort_values(by=["label_lower", "row_number", "bss"])

    # Group by label_lower and aggregate
    df = (
        df.groupby("label_lower")
        .agg(
            langs=(
                "lang",
                lambda x: ", ".join(x.sort_values().unique()),
            ),  # Create comma-separated list of unique languages
            datapoint_count=("datapoint_count", lambda x: x.fillna(0).sum()),
            # Force single True/False values to 1/0 and then sum.
            summary_count=("in_summary", lambda x: x.fillna(0).astype(int).sum()),
            unique_bss_count=("bss", pd.Series.nunique),
            min_row_number=("row_number", "min"),
            max_row_number=("row_number", "max"),
            bss_for_min_row=("bss", "first"),  # Assuming df is sorted by row_number within each group
            bss_for_max_row=("bss", "last"),  # Assuming df is sorted by row_number within each group
        )
        .reset_index()
        .rename(columns={"label_lower": "label"})
    )

    df = df.sort_values(by=["min_row_number", "label_lower", "bss_for_min_row", "bss_for_max_row"])

    # Add a translation of the label
    translator = Translator()

    def translate_label(label, langs):
        """
        Use the Google Translate API to translate the label from <lang> to English
        """
        langs = [lang.strip() for lang in langs.split(",") if lang.strip() != "en"]
        if not langs:
            return label
        try:
            translation = translator.translate(label, dest="en", src=langs[0])
            return translation.text
        except Exception:
            return ""

    df["translation"] = df.apply(lambda x: translate_label(x["label"], x["langs"]), axis="columns")

    # Find the assocatiated label metadata for these labels
    if label_type in ActivityLabel.LivelihoodActivityType.values:
        queryset = (
            ActivityLabel.objects.filter(activity_type=label_type)
            # Rename the label column to make it easy to merge the dataframes
            .annotate(label=F("activity_label")).values(
                "label",
                "activity_type",
                "status",
                "is_start",
                "strategy_type",
                "attribute",
                "product__common_name_en",
                "unit_of_measure_id",
                "currency_id",
                "season",
                "additional_identifier",
                "notes",
            )
        )
    elif label_type == "WealthCharacteristic":
        queryset = (
            WealthCharacteristicLabel.objects.all()
            # Rename the label column to make it easy to merge the dataframes
            .annotate(label=F("wealth_characteristic_label")).values(
                "status",
                "label",
                "wealth_characteristic_id",
                "product__common_name_en",
                "unit_of_measure_id",
                "notes",
            )
        )
    label_metadata_df = pd.DataFrame.from_records(queryset)

    # Merge the label metadata into the dataframe
    if not label_metadata_df.empty:
        df = df.merge(label_metadata_df, left_on="label", right_on="label", how="left")

    # Rename the columns to match what we need in the GSheet when we run jobs.metadata.load_all_metadata
    df = df.rename(
        columns={
            "label": "wealth_characteristic_label" if label_type == "WealthCharacteristic" else "activity_label",
            "product__common_name_en": "product_name",
        }
    )

    return Output(
        df,
        metadata={
            "num_labels": len(df),
            "num_datapoints": int(df["datapoint_count"].sum()),
            "num_summaries": int(df["summary_count"].sum()),
            "preview": MetadataValue.md(df.sample(config.preview_rows).to_markdown()),
            "datapoint_preview": MetadataValue.md(
                df[df["datapoint_count"] > 0].sample(config.preview_rows).to_markdown()
            ),
        },
    )
