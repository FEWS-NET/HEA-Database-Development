"""
Dagster assets related to Wealth Group Characteristics, read from the 'WB' worksheet in a BSS.

An example of relevant rows from the worksheet:
    |     | A                                              | B                   | C                                          | D                                          | E                                          | F                                          | G                |
    |----:|:-----------------------------------------------|:--------------------|:-------------------------------------------|:-------------------------------------------|:-------------------------------------------|:-------------------------------------------|:-----------------|
    |   1 | MALAWI HEA BASELINES 2015                      | Southern Lakeshore  | Southern Lakeshore                         |                                            |                                            |                                            |                  |
    |   2 |                                                |                     | Community interviews                       |                                            |                                            |                                            |                  |
    |   3 | WEALTH GROUP                                   |                     |                                            |                                            |                                            |                                            |                  |
    |   4 | District                                       | Salima and Mangochi | Salima                                     | Salima                                     | Salima                                     | Salima                                     | Dedza            |
    |   5 | Village                                        |                     | Mtika                                      | Pemba                                      | Ndembo                                     | Makanjira                                  | Kasakala         |
    |   6 | Interview number:                              |                     | 1                                          | 2                                          | 3                                          | 4                                          | 5                |
    |   7 | Interviewers                                   |                     | Kandiwo, Ethel, Fyawupi, Daniel, Chipiliro | Kandiwo, Ethel, Fyawupi, Daniel, Chipiliro | Kandiwo, Ethel, Fyawupi, Chipiliro, Daniel | Kandiwo, Ethel, Fyawupi, Chipiliro, Daniel | Chipiliro, Imran |
    |   8 | Wealth characteristics                         |                     |                                            |                                            |                                            |                                            |                  |
    |   9 | Wealth breakdown (% of households)             | VP                  | 36                                         | 35                                         | 52                                         | 43                                         | 41               |
    |  10 |                                                | P                   | 27                                         | 29                                         | 16                                         | 29                                         | 34               |
    |  11 |                                                | M                   | 29                                         | 23                                         | 23                                         | 20                                         | 18               |
    |  12 |                                                | B/O                 | 8                                          | 13                                         | 9                                          | 8                                          | 7                |
    |  13 |                                                |                     | 100                                        | 100                                        | 100                                        | 100                                        | 100              |
    |  14 | HH size                                        | VP                  | 8                                          | 8                                          | 7                                          | 8                                          | 7                |
    |  15 |                                                | P                   | 9                                          | 8                                          | 7                                          | 7                                          | 7                |
    |  16 |                                                | M                   | 10                                         | 5                                          | 7                                          | 6                                          | 7                |
    |  17 |                                                | B/O                 | 15                                         | 10                                         | 7                                          | 6                                          | 7                |
    |  18 | Land area owned (acres)                        | VP                  | 1.4                                        | 0.75                                       | 0.5                                        | 2                                          | 1                |
    |  19 |                                                | P                   | 1.4                                        | 0.75                                       | 1                                          | 2                                          | 1                |
    |  20 |                                                | M                   | 1.4                                        | 0.75                                       | 1.5                                        | 2                                          | 1                |
    |  21 |                                                | B/O                 | 1.4                                        | 0.75                                       | 2                                          | 2                                          | 2                |
    |  66 | Cattle: Oxen number owned                      | VP                  | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  67 |                                                | P                   | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  68 |                                                | M                   | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  69 |                                                | B/O                 | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  70 | Cattle: total owned at start of year           | VP                  | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  71 | adult females                                  | VP                  | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  72 | no.born during year                            | VP                  |                                            |                                            |                                            |                                            |                  |
    |  73 | no. sold                                       | VP                  |                                            |                                            |                                            |                                            |                  |
    |  74 | no. slaughtered                                | VP                  |                                            |                                            |                                            |                                            |                  |
    |  75 | no. died                                       | VP                  |                                            |                                            |                                            |                                            |                  |
    |  76 | no. bought                                     | VP                  |                                            |                                            |                                            |                                            |                  |
    |  77 | no. at end of reference year                   | VP                  |                                            |                                            |                                            |                                            |                  |
    |  78 | Cattle: total owned at start of year           | P                   | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  79 | adult females                                  | P                   | 0                                          | 0                                          | 0                                          | 0                                          | 0                |
    |  80 | no.born during year                            | P                   |                                            |                                            |                                            |                                            |                  |
    |  81 | no. sold                                       | P                   |                                            |                                            |                                            |                                            |                  |
    |  82 | no. slaughtered                                | P                   |                                            |                                            |                                            |                                            |                  |
    |  83 | no. died                                       | P                   |                                            |                                            |                                            |                                            |                  |
    |  84 | no. bought                                     | P                   |                                            |                                            |                                            |                                            |                  |
    |  85 | no. at end of reference year                   | P                   |                                            |                                            |                                            |                                            |                  |
    |  86 | Cattle: total owned at start of year           | M                   | 0                                          | 0                                          | 3                                          | 5                                          | 0                |
    |  87 | adult females                                  | M                   | 0                                          | 0                                          | 2                                          | 3                                          | 0                |
    |  88 | no.born during year                            | M                   |                                            |                                            |                                            |                                            |                  |
    |  89 | no. sold                                       | M                   |                                            |                                            |                                            |                                            |                  |
    |  90 | no. slaughtered                                | M                   |                                            |                                            |                                            |                                            |                  |
    |  91 | no. died                                       | M                   |                                            |                                            |                                            |                                            |                  |
    |  92 | no. bought                                     | M                   |                                            |                                            |                                            |                                            |                  |
    |  93 | no. at end of reference year                   | M                   |                                            |                                            |                                            |                                            |                  |
    |  94 | Cattle: total owned at start of year           | B/O                 | 15                                         | 9                                          | 30                                         | 18                                         | 3                |
    |  95 | adult females                                  | B/O                 | 13                                         | 6                                          | 24                                         | 14                                         | 2                |
    |  96 | no.born during year                            | B/O                 |                                            |                                            |                                            |                                            |                  |
    |  97 | no. sold                                       | B/O                 |                                            |                                            |                                            |                                            |                  |
    |  98 | no. slaughtered                                | B/O                 |                                            |                                            |                                            |                                            |                  |
    |  99 | no. died                                       | B/O                 |                                            |                                            |                                            |                                            |                  |
    | 100 | no. bought                                     | B/O                 |                                            |                                            |                                            |                                            |                  |
    | 101 | no. at end of reference year                   | B/O                 |                                            |                                            |                                            |                                            |                  |
    | 170 | Chicken number owned                           | VP                  | 6                                          | 8                                          | 1                                          | 3                                          | 3                |
    | 171 |                                                | P                   | 9                                          | 10                                         | 4                                          | 10                                         | 4                |
    | 172 |                                                | M                   | 13                                         | 10                                         | 10                                         | 30                                         | 6                |
    | 173 |                                                | B/O                 | 25                                         | 25                                         | 25                                         | 40                                         | 7                |
    | 186 | Main food crops                                | VP                  | 2                                          | 2                                          | 2                                          | 2                                          | 2                |
    | 187 |                                                | P                   | 2                                          | 2                                          | 2                                          | 2                                          | 2                |
    | 188 |                                                | M                   | 3                                          | 3                                          | 3                                          | 3                                          | 3                |
    | 189 |                                                | B/O                 | 3                                          | 3                                          | 3                                          | 3                                          | 3                |
    | 190 | Main cash crops                                | VP                  | 1                                          | 3                                          | 2                                          | 1                                          | 1                |
    | 191 |                                                | P                   | 2                                          | 3                                          | 2                                          | 1                                          | 1                |
    | 192 |                                                | M                   | 2                                          | 3                                          | 2                                          | 2                                          | 2                |
    | 193 |                                                | B/O                 | 2                                          | 3                                          | 2                                          | 2                                          | 2                |
    | 194 | Main source of cash income 1st                 | VP                  | Casual labour                              | Casual labour                              | Agric Labour                               | Agricultural Labour                        | Agric ganyu      |
    | 195 |                                                | P                   | agricultural labour                        | Casual labour                              | Handcrafts                                 | Agricultural Labour                        | Kabanza,         |
    | 196 |                                                | M                   | agricultural labour                        | small businesses                           | Crop sales                                 | Crop sales                                 | Kabaza business, |
    | 197 |                                                | B/O                 | Businesses                                 | crop sales                                 | Crop sales                                 | Fish sales                                 | Small business,  |
"""  # NOQA: E501

import itertools
import json
import os

import django
import pandas as pd
from dagster import AssetExecutionContext, MetadataValue, Output, asset
from openpyxl.utils import get_column_letter

from ..configs import BSSMetadataConfig
from ..partitions import bss_instances_partitions_def
from ..utils import get_index, prepare_lookup, verbose_pivot
from .base import (
    get_all_bss_labels_dataframe,
    get_bss_dataframe,
    get_bss_label_dataframe,
    get_summary_bss_label_dataframe,
)

# set the default Django settings module
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "hea.settings.production")

# Configure Django with our custom settings before importing any Django classes
django.setup()

from baseline.models import (  # NOQA: E402
    LivelihoodZoneBaseline,
    WealthGroupCharacteristicValue,
)
from metadata.lookups import WealthGroupCategoryLookup  # NOQA: E402
from metadata.models import WealthCharacteristicLabel  # NOQA: E402

# Indexes of header rows in the Data3 dataframe (wealth_group_category, district, village)
HEADER_ROWS = [3, 4, 5]


@asset(partitions_def=bss_instances_partitions_def)
def wealth_characteristic_dataframe(config: BSSMetadataConfig, corrected_files) -> Output[pd.DataFrame]:
    """
    DataFrame of Wealth Group Characteristic Values from a BSS
    """
    return get_bss_dataframe(
        config,
        corrected_files,
        "WB",
        start_strings=["Wealth characteristics", "Caractéristiques socio-économiques", "Caractéristiques de richesse"],
        end_strings=["Informations sur les équidés"],
        header_rows=HEADER_ROWS,
        # The final three relevant columns are Summary/From/To in Row 4. Range/Interval will be in the cell above
        # From (i.e. in Row 3) so force two additional summary columns.
        num_summary_cols=2,
    )


@asset(partitions_def=bss_instances_partitions_def)
def wealth_characteristic_label_dataframe(
    context: AssetExecutionContext,
    config: BSSMetadataConfig,
    wealth_characteristic_dataframe,
) -> Output[pd.DataFrame]:
    """
    Dataframe of Wealth Group Characteristic Label References
    """
    return get_bss_label_dataframe(
        context, config, wealth_characteristic_dataframe, "wealth_characteristic_dataframe", len(HEADER_ROWS)
    )


@asset(io_manager_key="dataframe_csv_io_manager")
def all_wealth_characteristic_labels_dataframe(
    config: BSSMetadataConfig, wealth_characteristic_label_dataframe: dict[str, pd.DataFrame]
) -> Output[pd.DataFrame]:
    """
    Combined dataframe of the Wealth Group Characteristic labels in use across all BSSs.
    """
    return get_all_bss_labels_dataframe(config, wealth_characteristic_label_dataframe)


@asset(io_manager_key="dataframe_csv_io_manager")
def summary_wealth_characteristic_labels_dataframe(
    config: BSSMetadataConfig, all_wealth_characteristic_labels_dataframe: pd.DataFrame
) -> Output[pd.DataFrame]:
    """
    Summary of the Wealth Group Characteristic labels in use across all BSSs.
    """
    return get_summary_bss_label_dataframe(config, all_wealth_characteristic_labels_dataframe)


@asset(partitions_def=bss_instances_partitions_def, io_manager_key="json_io_manager")
def wealth_characteristic_instances(
    context: AssetExecutionContext,
    config: BSSMetadataConfig,
    wealth_characteristic_dataframe,
) -> Output[dict]:
    """
    WealthGroup and WealthGroupCharacteristicValue instances extracted from the BSS.
    """
    # Find the metadata for this BSS
    partition_key = context.asset_partition_key_for_output()
    livelihood_zone_baseline = LivelihoodZoneBaseline.objects.get_by_natural_key(*partition_key.split("~")[1:])

    df = wealth_characteristic_dataframe
    num_header_rows = 3  # wealth group category, district, village

    # Prepare the lookups, so they cache the individual results
    wealthgroupcategorylookup = WealthGroupCategoryLookup()
    label_map = {
        instance.pop("wealth_characteristic_label").lower(): instance
        for instance in WealthCharacteristicLabel.objects.values(
            "wealth_characteristic_label",
            "wealth_characteristic_id",
            "product_id",
            "unit_of_measure_id",
        )
    }
    context.log.info("Loaded %d Wealth Characteristic Labels", len(label_map))

    # Create the Wealth Groups from the combination of Wealth Group Categories from column B and Communities from
    # Rows 4 and 5. The Communities are repeated for each Wealth Group Category, so only use the values until the
    # first Wealth Group Category.
    # Find the last Community from the Community (Form 3) Interviews, which don't have a Wealth Group Category
    row3_categories = df.loc[3, "B":].unique()
    row3_categories = row3_categories[row3_categories != ""]
    last_form3_col = get_index(row3_categories, df.loc[3, "C":], offset=-1)
    df.loc[4] = df.loc[4].str.strip().replace("", pd.NA)
    df.loc[5] = df.loc[5].str.strip().replace("", pd.NA)
    form3_full_names = [
        ", ".join(df.loc[4:5, column].sort_index(ascending=False).dropna())
        for column in df.loc[4:5, "C":last_form3_col].dropna(how="all", axis="columns")
    ]
    wealth_group_categories = [
        wealth_group_category for wealth_group_category in df.loc[:, "B"].replace("", pd.NA).dropna().unique()
    ]
    wealth_group_df = pd.DataFrame(
        list(itertools.product(wealth_group_categories, form3_full_names)),
        columns=["wealth_group_category", "full_name"],
    )
    # There is also a set of Summary Wealth Groups that don't contain a Community
    wealth_group_df = pd.concat(
        [wealth_group_df, pd.DataFrame({"wealth_group_category": wealth_group_categories, "full_name": ""})]
    ).reset_index(drop=True)
    # Lookup the wealth group category
    wealth_group_df = WealthGroupCategoryLookup().do_lookup(
        wealth_group_df, "wealth_group_category", "wealth_group_category"
    )
    # Add the natural key for the Livelihood Zone Baseline to the Wealth Groups
    wealth_group_df["livelihood_zone_baseline"] = [
        [livelihood_zone_baseline.livelihood_zone_id, livelihood_zone_baseline.reference_year_end_date.isoformat()]
    ] * len(wealth_group_df)
    # Add the natural key for the Community to the Wealth Groups
    wealth_group_df["community"] = wealth_group_df[["livelihood_zone_baseline", "full_name"]].apply(
        lambda x: x.iloc[0] + [x.iloc[1]] if x.iloc[1] else None, axis="columns"
    )

    # Build a list of the Community Full Name for each column, based on the values from Rows 4 and 5. For rows that
    # don't have a value in either row 4 or row 5, and the last 3 columns that are the Summary, return ""
    community_full_names = [
        (
            ", ".join(df.loc[4:5, column].sort_index(ascending=False).dropna())
            if df.loc[4:5, column].any() and column not in df.columns[-3:]
            else ""
        )
        for column in df.loc[4:5, "C":]
    ]

    # Build a list of the Wealth Group Categories for each column, based on the values from Row 3.
    wealth_group_categories = [
        wealthgroupcategorylookup.get(wealth_group_category) if wealth_group_category else ""
        for wealth_group_category in df.loc[3, "C":]
    ]

    # Prepare the label column for matching against the label_map
    prepared_labels = prepare_lookup(df["A"])

    # Check that we recognize all of the wealth characteristic labels
    allow_unrecognized_labels = True
    unrecognized_labels = (
        df.iloc[num_header_rows:][
            ~prepared_labels.iloc[num_header_rows:].isin(label_map) & (prepared_labels.iloc[num_header_rows:] != "")
        ]
        .groupby("A")
        .apply(lambda x: ", ".join(x.index.astype(str)))
    )
    if unrecognized_labels.empty:
        unrecognized_labels = pd.DataFrame(columns=["label", "rows"])
    else:
        unrecognized_labels = unrecognized_labels.reset_index()
        unrecognized_labels.columns = ["label", "rows"]
        message = "Unrecognized wealth characteristic labels:\n\n" + unrecognized_labels.to_markdown(index=False)
        if allow_unrecognized_labels:
            context.log.warning(message)
        else:
            raise ValueError(message)

    # Process the main part of the sheet to find the Wealth Group Characteristic Values

    # Regex matching is unreliable because of inconsistencies across BSSs from different countries. However, there are
    # a finite number of wealth characteristic labels (the value in column A) in use across the global set of BSSs.
    # Therefore, we map the strings to attributes directly.

    # Although the structure of this worksheet is not as complicated as the Data sheet, and we could build the fixture
    # using vector DataFrame operations, it is easier to maintain this code if it follows the same structure as the
    # `livelihood_activity_instances`. Therefore, we iterate over the rows rather than use vector operations.

    # Iterate over the rows
    wealth_group_characteristic_values = []
    for row in df.iloc[num_header_rows:].index:  # Ignore the Wealth Group header rows
        label = prepared_labels[row]
        if not label:
            # Ignore blank rows
            continue
        # Get the attributes, taking a copy so that we can pop() some of the attributes without altering the original
        attributes = label_map.get(label, {}).copy()
        if not any(attributes.values()):
            # Ignore rows that don't contain any relevant data (or which aren't in the label_map)
            continue
        # Lookup the Wealth Group Category from Column B
        wealth_group_category = wealthgroupcategorylookup.get(df.loc[row, "B"])
        if not wealth_group_category:
            # Ignore rows that don't contain a Wealth Group Category, which are typically calculated totals of the
            # per-Wealth Group Category values, e.g. MWKAS_30Sep15.xlsx has a row 13 that contains the sum of the
            # percentage of households data in rows 9-12.
            continue

        # Create the WealthGroupCharacteristic records
        if any(value for value in df.loc[row, "C":]):
            # Iterate over the value columns, from Column C to the the Summary Column.
            # We don't iterate over the last two columns because they contain the min_value and max_value that are
            # part of the Summary Wealth Characteristic Value rather than a separate Wealth Characteristic Value.
            for i, value in enumerate(df.loc[row, "C" : df.columns[-3]]):
                # Store the column to aid trouble-shooting.
                # We need col_index + 1 to get the letter, and the enumerate is already starting from col C
                column = get_column_letter(i + 3)
                try:
                    # Add find the reference_type:
                    # Wealth Group (Form 4) values will have a full name and a wealth group category from Row 3
                    if community_full_names[i] and wealth_group_categories[i]:
                        reference_type = WealthGroupCharacteristicValue.CharacteristicReference.WEALTH_GROUP
                    # Community (Form 3) values will have a full name from Rows 4 and 5, but no wealth group category
                    elif community_full_names[i]:
                        reference_type = WealthGroupCharacteristicValue.CharacteristicReference.COMMUNITY
                    # Summary values will not have full name or a wealth category, and will be in the last 3 columns
                    # Check for len(df.columns) -5 because the Summary col is 3rd from end, and i starts at Column C.
                    elif i == len(df.columns) - 5:
                        reference_type = WealthGroupCharacteristicValue.CharacteristicReference.SUMMARY
                    # There is no full name, and this isn't the summary, so we can ignore this column. This happens
                    # because there are typically blank columns in BSS between each wealth group category. For example,
                    # in MWKAS_30Sep15.xlsx they are in columns L, V, AF, AP and AZ
                    else:
                        reference_type = None

                    # Only store Wealth Group Characteristic Values that have a value (which might be zero).
                    # Ignore columns where we couldn't determine the reference_type. Even though these are
                    # supposed to be blank columns, they sometimes contain 0 values for some rows, particularly where
                    # the Wealth Characteristic Value for the adjacent columns is also 0.
                    # Ignore Form 4 Wealth Group Characteristic Values where the Value is for a different Wealth Group
                    # to the one being interviewed. I.e. if there is a Wealth Group Category on Row 3 it must match the
                    # Wealth Group Category in Column B.
                    if (
                        value != ""
                        and reference_type
                        and (not wealth_group_categories[i] or wealth_group_categories[i] == wealth_group_category)
                    ):
                        wealth_group_characteristic_value = attributes.copy()

                        # The natural key for the Wealth Group is made up of the Livelihood Zone Baseline, the
                        # Wealth Group Category from column B and the Community Full Name from Rows 4 and 5.
                        wealth_group_characteristic_value["wealth_group"] = [
                            livelihood_zone_baseline.livelihood_zone_id,
                            livelihood_zone_baseline.reference_year_end_date.isoformat(),
                            wealth_group_category,
                            community_full_names[i],
                        ]

                        wealth_group_characteristic_value["reference_type"] = reference_type

                        # The percentage of households should be stored as a number between 1 and 100,
                        # but may be stored in the BSS (particularly in the summary column) as a
                        # decimal fraction between 0 and 1, so correct those values
                        if (
                            wealth_group_characteristic_value["wealth_characteristic_id"] == "percentage of households"
                            and value != ""
                            and float(value) < 1
                        ):
                            value = float(value) * 100

                        wealth_group_characteristic_value["value"] = value

                        # If this is the summary, then also save the min and max values
                        if reference_type == WealthGroupCharacteristicValue.CharacteristicReference.SUMMARY:
                            min_value = df.loc[row, df.columns[-2]]
                            if min_value != "" and float(min_value) < 1:
                                min_value = float(min_value) * 100
                            max_value = df.loc[row, df.columns[-1]]
                            if max_value != "" and float(max_value) < 1:
                                max_value = float(max_value) * 100
                            wealth_group_characteristic_value["min_value"] = min_value
                            wealth_group_characteristic_value["max_value"] = max_value

                        # Save the column and row, to aid trouble-shooting
                        wealth_group_characteristic_value["bss_sheet"] = "WB"
                        wealth_group_characteristic_value["bss_column"] = column
                        wealth_group_characteristic_value["bss_row"] = row
                        wealth_group_characteristic_values.append(wealth_group_characteristic_value)
                except Exception as e:
                    raise RuntimeError(
                        "Unhandled error in %s processing cell 'WB'!%s%s" % (partition_key, column, row)
                    ) from e

    # Create a dataframe of the Wealth Group Characteristic Values so that we can extract the
    # percentage of households and average household size, and run additional validation.
    value_df = pd.DataFrame.from_records(wealth_group_characteristic_values)
    value_df["wealth_group_category"] = value_df["wealth_group"].apply(lambda wealth_group: wealth_group[2])
    value_df["full_name"] = value_df["wealth_group"].apply(lambda wealth_group: wealth_group[3])

    # Make sure that the names in the Wealth Group-level interviews (e.g. columns $M:$AZ) match
    # the names in the in the Community-level interviews (e.g. columns $C:$K) that were used to
    # create the Wealth Group records
    form3_full_names_df = pd.DataFrame(
        [(get_column_letter(i + 3), full_name) for i, full_name in enumerate(form3_full_names)],
        columns=["bss_column", "full_name"],
    )
    unmatched_full_names = value_df[
        pd.notna(value_df["full_name"].replace("", pd.NA))
        & ~value_df["full_name"].str.lower().isin(form3_full_names_df.full_name.str.lower())
        # The Wealth Group Interview columns may only have the "Village" name without the "District", so also check
        # them against just the first part of the actual full names.
        & ~value_df["full_name"]
        .str.lower()
        .isin(form3_full_names_df.full_name.str.lower().apply(lambda x: x.split(", ")[0]))
    ][["bss_column", "full_name"]].drop_duplicates()
    if not unmatched_full_names.empty:
        raise ValueError(
            "%s contains unmatched Community full_name values in Wealth Group interviews:\n%s\n\nExpected names:\n%s"
            % (
                partition_key,
                unmatched_full_names.to_markdown(index=False),
                form3_full_names_df.to_markdown(index=False),
            )
        )

    # Add the percentage of households and average household size to the wealth groups
    # Filter the Wealth Group Characteristic Values to just those attribute,
    # where the source is either the Wealth Group Interview or the Summary.
    # (The Community Interview values aren't used for the Wealth Group household size or percentage of households).
    extra_attributes_df = value_df[
        value_df["wealth_characteristic_id"].isin(["percentage of households", "household size"])
        & value_df["reference_type"].isin(
            [
                WealthGroupCharacteristicValue.CharacteristicReference.WEALTH_GROUP,
                WealthGroupCharacteristicValue.CharacteristicReference.SUMMARY,
            ]
        )
    ][["wealth_group_category", "full_name", "wealth_characteristic_id", "value"]]
    extra_attributes_df = verbose_pivot(
        extra_attributes_df,
        values="value",
        index=["wealth_group_category", "full_name"],
        columns="wealth_characteristic_id",
    )
    extra_attributes_df = extra_attributes_df.rename(
        columns={
            "percentage of households": "percentage_of_households",
            "household size": "average_household_size",
        }
    )
    wealth_group_df = pd.merge(
        wealth_group_df, extra_attributes_df, on=["full_name", "wealth_group_category"], how="left"
    )

    result = {
        "WealthGroup": wealth_group_df.to_dict(orient="records"),
        "WealthGroupCharacteristicValue": wealth_group_characteristic_values,
    }
    metadata = {
        "num_wealth_groups": len(wealth_group_df),
        "num_wealth_group_characteristic_values": len(wealth_group_characteristic_values),
        "num_unrecognized_labels": len(unrecognized_labels),
        "pct_rows_recognized": round(
            (
                1
                - len(df.iloc[num_header_rows:][df.iloc[num_header_rows:]["A"].isin(unrecognized_labels["label"])])
                / len(df.iloc[num_header_rows:])
            )
            * 100
        ),
        "preview": MetadataValue.md(f"```json\n{json.dumps(result, indent=4)}\n```"),
    }
    if not unrecognized_labels.empty:
        metadata["unrecognized_labels"] = MetadataValue.md(unrecognized_labels.to_markdown(index=False))

    return Output(
        result,
        metadata=metadata,
    )
